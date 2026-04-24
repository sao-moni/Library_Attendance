import re
import sys
import os
import json
import logging
import cv2
import time
import sqlite3
import pandas as pd
import shutil
import traceback
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QLabel, QPushButton, QVBoxLayout, QHBoxLayout,
    QComboBox, QProgressBar, QListWidget, QDialog, QLineEdit, QMessageBox, QFileDialog,
    QTabWidget, QTableWidget, QTableWidgetItem, QDateEdit, QCheckBox, QHeaderView, QGridLayout,
    QListWidgetItem, QRadioButton, QAction, QInputDialog, QMenuBar, QProgressDialog, QFrame, QTextBrowser,
    QSplitter, QSizePolicy, QScrollArea, QGroupBox, QGraphicsDropShadowEffect, QSpinBox
)
from PyQt5.QtGui import QImage, QPixmap, QIcon, QFont, QPainter, QColor, QLinearGradient, QBrush, QPen, QFontDatabase, QPainterPath, QFontMetrics
from PyQt5.QtCore import Qt, QTimer, QThread, pyqtSignal, QDate, QUrl, QRect, QSize, QPoint, QPropertyAnimation, QEasingCurve, pyqtProperty
from PyQt5.QtMultimedia import QMediaPlayer, QMediaContent
from PyQt5.QtPrintSupport import QPrinter, QPrintDialog
from PyQt5.QtChart import QChart, QChartView, QBarSeries, QBarSet, QBarCategoryAxis, QValueAxis, QPieSeries, QLineSeries
import qrcode
from PyQt5.QtGui import QDesktopServices
import io
from io import BytesIO
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from threading import Lock
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet
import datetime
from datetime import date, timedelta
from itertools import takewhile
from PIL import Image as PILImage, ImageDraw, ImageFont
import numpy as np

# Check for module shadowing
if os.path.exists(os.path.join(os.path.dirname(__file__), 'datetime.py')):
    raise ImportError(
        "A file named 'datetime.py' exists in the project directory, shadowing the standard library module. Rename or remove it.")

# --- Constants for Settings ---
SETTINGS_FILE = "settings.json"
DEFAULT_SETTINGS = {
    "sound_enabled": True,
    "camera_index": 0,
    "default_study_year": "2025-2026",
    "qr_save_dir": "library_data/qr_codes",
}


# Configuration Constants
THEME_BG_COLOR = "#F5F6F5"
TEXT_COLOR = "#2D3436"
BUTTON_COLOR = "#0984E3"
BUTTON_HOVER_COLOR = "#74B9FF"
BUTTON_TEXT_COLOR = "#FFFFFF"
VIDEO_BG_COLOR = "#D8DEE9"
USER_DATA_FILE = "library_users.json"
LOG_DATABASE_FILE = "library_logs.db"
QR_CODE_SAVE_DIR = "library_data/qr_codes"
APP_TITLE = "Smart Library Analytics Hub"
INVAILD_SOUND = "asset/sound/invaild.mp3"
CHECK_IN_SOUND = "asset/sound/checkin.mp3"
CHECK_OUT_SOUND = "asset/sound/checkout.mp3"
FIXED_FONT_FAMILY = "Khmer OS System"  # Fixed font for the entire application
FIXED_FONT_SIZE = 10  # Fixed font size in points
UI_FONT_FAMILY = FIXED_FONT_FAMILY
FONT_FAMILY = FIXED_FONT_FAMILY
BASE_FONT_SIZE = FIXED_FONT_SIZE  # Kept for reference but not read from settings
LARGE_FONT_SIZE = 12
GROUP_SPACING = 12
DIALOG_PADDING = 15
SCAN_DELAY_MS = 1000
CAMERA_FPS = 30
CARD_BG_COLOR = "#0984E3"
CARD_TEXT_COLOR = "#FFFFFF"
DEFAULT_STUDY_YEAR = "2025-2026"
VIDEO_FIXED_SIZE = (640, 480)

# Logging Configuration
logging.basicConfig(filename='library_access.log', level=logging.DEBUG,
                    format='%(asctime)s - %(levelname)s - %(message)s')
os.makedirs(QR_CODE_SAVE_DIR, exist_ok=True)

# Custom Widget: Animated Progress Ring for Dashboard


class ProgressRing(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._value = 0
        self._text = ""
        self.setMinimumSize(150, 150)

    def setValue(self, value):
        self._value = max(0, min(100, value))
        self.update()

    def setText(self, text):
        self._text = text
        self.update()

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)

        rect = self.rect()
        center = rect.center()
        radius = min(rect.width(), rect.height()) // 2 - 10

        # Draw background ring
        bg_color = QColor("#E0E0E0")
        painter.setPen(QPen(bg_color, 8))
        painter.drawEllipse(center, radius, radius)

        # Draw progress arc
        if self._value > 0:
            progress_color = QColor(BUTTON_COLOR)
            painter.setPen(QPen(progress_color, 8))
            start_angle = 90 * 16  # Start from top
            # Convert percentage to degrees
            span_angle = -int(self._value * 3.6) * 16
            painter.drawArc(center.x() - radius, center.y() - radius,
                            radius * 2, radius * 2, start_angle, span_angle)

        # Draw text
        painter.setPen(QColor(TEXT_COLOR))
        font = QFont(UI_FONT_FAMILY, 16, QFont.Bold)
        painter.setFont(font)
        painter.drawText(rect, Qt.AlignCenter, self._text)

# Custom Widget: Card with Shadow for Dashboard


class DashboardCard(QWidget):
    def __init__(self, title="", parent=None):
        super().__init__(parent)
        self.title = title
        self.setup_ui()

    def setup_ui(self):
        self.setStyleSheet("""
            DashboardCard {
                background-color: white;
                border-radius: 12px;
                border: 1px solid #E0E0E0;
            }
        """)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(15, 15, 15, 15)
        layout.setSpacing(10)

        # Title
        title_label = QLabel(self.title)
        title_label.setFont(QFont(UI_FONT_FAMILY, 12, QFont.Bold))
        title_label.setStyleSheet(f"color: {TEXT_COLOR};")
        layout.addWidget(title_label)

        # Content placeholder
        self.content_layout = QVBoxLayout()
        layout.addLayout(self.content_layout)

        # Add shadow effect
        shadow = QGraphicsDropShadowEffect(self)
        shadow.setBlurRadius(15)
        shadow.setXOffset(0)
        shadow.setYOffset(2)
        shadow.setColor(QColor(0, 0, 0, 60))
        self.setGraphicsEffect(shadow)

    def addWidget(self, widget):
        self.content_layout.addWidget(widget)

# User Manager Class


class UserManager:
    def __init__(self, data_file):
        self.data_file = data_file
        self.users_cache = None
        self.lock = Lock()

    def load_users(self):
        with self.lock:
            if self.users_cache is not None:
                return self.users_cache
            try:
                if not os.path.exists(self.data_file):
                    with open(self.data_file, "w") as f:
                        json.dump([], f)
                    self.users_cache = []
                with open(self.data_file, "r") as file:
                    self.users_cache = json.load(file)
                return self.users_cache
            except (json.JSONDecodeError, IOError) as e:
                logging.error(
                    f"Failed to load users from {self.data_file}: {e}")
                self.users_cache = []
                return self.users_cache

    def save_users(self, users):
        with self.lock:
            try:
                with open(self.data_file, "w") as file:
                    json.dump(users, file, indent=4)
                self.users_cache = users
            except IOError as e:
                logging.error(f"Failed to save users to {self.data_file}: {e}")

    def add_user(self, user_data):
        users = self.load_users()
        user_id = user_data.get('id')
        if user_id and any(user['id'] == user_id for user in users):
            raise ValueError(f"User ID {user_id} already exists.")
        if not user_id:
            user_id = max([user['id'] for user in users], default=0) + 1
            user_data['id'] = user_id
        users.append(user_data)
        self.save_users(users)
        return user_id

    def delete_user(self, user_id):
        users = self.load_users()
        initial_len = len(users)
        users = [user for user in users if user['id'] != user_id]
        if len(users) < initial_len:
            self.save_users(users)
            return True
        return False

    def get_user_by_id(self, user_id):
        return next((user for user in self.load_users() if user['id'] == user_id), None)

    def get_all_users(self):
        return self.load_users()

    def get_classes(self):
        return sorted(set(user.get('class', '') for user in self.load_users() if user.get('class', '')))

# Log Manager Class


class LogManager:
    def __init__(self, db_file):
        self.db_file = db_file
        self.conn = sqlite3.connect(self.db_file, check_same_thread=False)
        self.lock = Lock()
        self.today_cache = None
        self.today_date = None
        self.initialize_database()

    def initialize_database(self):
        with self.lock:
            cursor = self.conn.cursor()
            cursor.execute('''CREATE TABLE IF NOT EXISTS logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                event TEXT NOT NULL,
                timestamp TEXT NOT NULL)''')
            self.conn.commit()

    def log_event(self, user_id, event):
        with self.lock:
            try:
                cursor = self.conn.cursor()
                timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                cursor.execute("INSERT INTO logs (user_id, event, timestamp) VALUES (?, ?, ?)",
                               (user_id, event, timestamp))
                self.conn.commit()
                self.today_cache = None
                self.today_date = None
                logging.debug(
                    f"Logged {event} for user {user_id} at {timestamp}")
                return timestamp
            except sqlite3.Error as e:
                logging.error(f"Failed to log event for user {user_id}: {e}")
                return None

    def get_today_logs(self):
        today = datetime.datetime.now().date()
        if self.today_date != today or self.today_cache is None:
            with self.lock:
                cursor = self.conn.cursor()
                cursor.execute("SELECT user_id, event, timestamp FROM logs WHERE date(timestamp) = ? ORDER BY timestamp ASC",
                               (today.strftime('%Y-%m-%d'),))
                self.today_cache = cursor.fetchall()
                self.today_date = today
        return self.today_cache

    def get_detailed_log_data_for_date_range(self, start_date, end_date):
        try:
            user_manager = UserManager(USER_DATA_FILE)
            users = user_manager.get_all_users()
            if not users:
                return []
            log_data = []
            with self.lock:
                cursor = self.conn.cursor()
                for user in users:
                    cursor.execute(
                        "SELECT event, timestamp FROM logs WHERE user_id = ? AND timestamp BETWEEN ? AND ? ORDER BY timestamp ASC",
                        (user['id'], start_date.strftime('%Y-%m-%d 00:00:00'), end_date.strftime('%Y-%m-%d 23:59:59')))
                    user_logs = pd.DataFrame(cursor.fetchall(), columns=[
                                             'event', 'timestamp'])
                    if not user_logs.empty:
                        log_data.append({
                            'user_id': user['id'], 'first_name': user['first_name'], 'last_name': user['last_name'],
                            'gender': user['gender'], 'class': user['class'], 'role': user['role'],
                            'study_year': user.get('study_year', DEFAULT_STUDY_YEAR), 'logs': user_logs
                        })
            return log_data
        except sqlite3.Error as e:
            logging.error(f"Failed to fetch logs: {e}")
            return []

    def get_last_event_for_user(self, user_id):
        today_logs = self.get_today_logs()
        user_logs = [log for log in today_logs if log[0] == user_id]
        return user_logs[-1][1] if user_logs else None

    def get_recent_events(self, limit=10):
        with self.lock:
            try:
                cursor = self.conn.cursor()
                cursor.execute(
                    "SELECT user_id, event, timestamp FROM logs ORDER BY timestamp DESC LIMIT ?", (limit,))
                return cursor.fetchall()
            except sqlite3.Error as e:
                logging.error(f"Failed to get recent events: {e}")
                return []

    def get_daily_stats(self, date):
        log_data = self.get_detailed_log_data_for_date_range(date, date)

        # Create a set of unique user IDs that have at least one 'Check-in'
        checked_in_users = {
            user['user_id']
            for user in log_data
            if not user['logs'][user['logs']['event'] == 'Check-in'].empty
        }

        # Create a set of unique female user IDs from the checked-in users
        female_checked_in_users = {
            user['user_id']
            for user in log_data
            if user['user_id'] in checked_in_users and user['gender'] == 'ស្រី'
        }

        return len(checked_in_users), len(female_checked_in_users)

    def get_grade_level_current_status(self, date):
        logs = self.get_detailed_log_data_for_date_range(date, date)
        grade_status = {grade: {'total_checked_in': 0,
                                'female_checked_in': 0} for grade in range(7, 13)}
        user_manager = UserManager(USER_DATA_FILE)
        users = {user['id']: user for user in user_manager.get_all_users()}
        for user_data in logs:
            user_id = user_data['user_id']
            user = users.get(user_id)
            if not user or 'class' not in user or not user['class']:
                logging.debug(
                    f"Skipping user {user_id} with missing or invalid class")
                continue
            class_name = user['class'].strip()
            # Extract grade using takewhile
            grade_str = ''.join(takewhile(str.isdigit, class_name))
            try:
                grade = int(grade_str) if grade_str else None
                if grade and 7 <= grade <= 12:
                    # Count check-ins for the day
                    check_ins = sum(
                        1 for _, log in user_data['logs'].iterrows() if log['event'] == 'Check-in')
                    if check_ins > 0:  # User has at least one check-in for the day
                        grade_status[grade]['total_checked_in'] += 1
                        if user['gender'].lower() == 'ស្រី':
                            grade_status[grade]['female_checked_in'] += 1
                else:
                    logging.debug(
                        f"Grade {grade} out of range or invalid for user {user_id} in class {class_name}")
            except ValueError:
                logging.debug(
                    f"Invalid grade format for user {user_id} in class {class_name}")
        return grade_status

    def get_total_check_ins(self, date):
        with self.lock:
            try:
                cursor = self.conn.cursor()
                cursor.execute("SELECT COUNT(*) FROM logs WHERE event = 'Check-in' AND date(timestamp) = ?",
                               (date.strftime('%Y-%m-%d'),))
                result = cursor.fetchone()
                return result[0] if result else 0
            except sqlite3.Error as e:
                logging.error(f"Failed to get total check-ins for {date}: {e}")
                return 0

    def get_unique_check_ins(self, date):
        with self.lock:
            try:
                cursor = self.conn.cursor()
                cursor.execute("SELECT COUNT(DISTINCT user_id) FROM logs WHERE event = 'Check-in' AND date(timestamp) = ?",
                               (date.strftime('%Y-%m-%d'),))
                result = cursor.fetchone()
                return result[0] if result else 0
            except sqlite3.Error as e:
                logging.error(
                    f"Failed to get unique check-ins for {date}: {e}")
                return 0

    def clear_logs(self):
        with self.lock:
            try:
                cursor = self.conn.cursor()
                cursor.execute("DELETE FROM logs")
                self.conn.commit()
                self.today_cache = None
                logging.info("All logs cleared.")
            except sqlite3.Error as e:
                logging.error(f"Failed to clear logs: {e}")
                raise

    def close(self):
        self.conn.close()

    def get_top_users_by_checkins(self, limit=10, start_date=None, end_date=None):
        """
        Retrieves the top N users based on their 'Check-in' event count within a date range.
        :param limit: The maximum number of top users to return (default 10).
        :param start_date: datetime.date object for the start of the period (inclusive).
        :param end_date: datetime.date object for the end of the period (inclusive).
        :return: A list of (user_id, checkin_count) tuples, sorted by count descending.
        """
        with self.lock:
            try:
                cursor = self.conn.cursor()
                query = """
                    SELECT user_id, COUNT(*) as checkin_count
                    FROM logs
                    WHERE event = 'Check-in'
                """
                params = []
                if start_date and end_date:
                    query += " AND date(timestamp) BETWEEN ? AND ?"
                    params.append(start_date.strftime('%Y-%m-%d'))
                    params.append(end_date.strftime('%Y-%m-%d'))
                elif start_date:  # If only start date is given, assume for that specific day
                    query += " AND date(timestamp) = ?"
                    params.append(start_date.strftime('%Y-%m-%d'))
                elif end_date:  # If only end date is given, assume up to that day
                    query += " AND date(timestamp) <= ?"
                    params.append(end_date.strftime('%Y-%m-%d'))
                query += " GROUP BY user_id ORDER BY checkin_count DESC LIMIT ?"
                params.append(limit)
                cursor.execute(query, tuple(params))
                return cursor.fetchall()
            except sqlite3.Error as e:
                logging.error(f"Failed to get top users by check-ins: {e}")
                return []

    def get_monthly_trend_data(self, year):
        """Fetches monthly check-in data for a given year."""
        with self.lock:
            try:
                cursor = self.conn.cursor()
                monthly_data = [0] * 12
                for month in range(1, 13):
                    start_date = f"{year}-{month:02d}-01"
                    if month == 12:
                        end_date = f"{year+1}-01-01"
                    else:
                        end_date = f"{year}-{month+1:02d}-01"
                    cursor.execute("SELECT COUNT(*) FROM logs WHERE event = 'Check-in' AND timestamp >= ? AND timestamp < ?",
                                   (start_date, end_date))
                    result = cursor.fetchone()
                    monthly_data[month-1] = result[0] if result else 0
                return monthly_data
            except sqlite3.Error as e:
                logging.error(
                    f"Failed to get monthly trend data for {year}: {e}")
                return [0] * 12

# Main Application Class


class QRCodeApp(QMainWindow):
    def __init__(self, user_manager, log_manager):
        super().__init__()
        self.user_manager = user_manager
        self.log_manager = log_manager
        self.settings = self.load_settings()
        self.settings["stats_date"] = datetime.datetime.now().date()

        self.is_camera_active = False
        self.current_user_id = None
        self.camera_thread = None
        self.scan_timer = QTimer(self)
        self.scan_timer.timeout.connect(self.process_scanning)
        self.scan_interval = 100
        self.last_scanned_code = None
        self.scan_cooldown = False
        self.media_player = QMediaPlayer(self)
        self.available_cameras = []

        # Initialize UI components that are shared across tabs
        self.recent_list = QListWidget()

        self.init_ui()
        self.check_sound_files()
        self.check_camera_availability()
        self.statusBar()

    def styled_label(self, text, color=TEXT_COLOR, font_size=None, bold=False, italic=False):
        """Creates a QLabel with consistent styling and font for Khmer support."""
        label = QLabel(text)

        # Determine font size, default to global fixed size
        if font_size is None:
            font_size = FIXED_FONT_SIZE

        font = QFont(UI_FONT_FAMILY, font_size)
        font.setBold(bold)
        font.setItalic(italic)

        label.setFont(font)
        label.setStyleSheet(
            f"color: {color}; background-color: transparent; padding: 2px;")
        return label

    def load_settings(self):
        """Loads settings from the JSON file, or creates it with defaults."""
        try:
            if not os.path.exists(SETTINGS_FILE):
                with open(SETTINGS_FILE, "w") as f:
                    json.dump(DEFAULT_SETTINGS, f, indent=4)
                return DEFAULT_SETTINGS
            with open(SETTINGS_FILE, "r") as f:
                settings = json.load(f)
                # Ensure all default keys are present
                for key, value in DEFAULT_SETTINGS.items():
                    if key not in settings:
                        settings[key] = value
                return settings
        except (IOError, json.JSONDecodeError) as e:
            logging.error(f"Failed to load settings from {SETTINGS_FILE}: {e}")
            return DEFAULT_SETTINGS  # Fallback to defaults

# In class QRCodeApp
    def save_settings(self, new_settings_data):
        """Saves the settings dictionary to the JSON file and updates the in-memory settings."""
        try:
            # 1. Update the application's current settings in-memory
            self.settings.update(new_settings_data)

            # 2. Create a clean copy of the NOW-UPDATED settings to save to disk
            settings_to_save = self.settings.copy()

            # 3. Remove any runtime-only keys before writing to the file
            if "stats_date" in settings_to_save:
                del settings_to_save["stats_date"]

            # 4. Write the clean dictionary to the JSON file
            with open(SETTINGS_FILE, "w") as f:
                json.dump(settings_to_save, f, indent=4)

        except IOError as e:
            logging.error(f"Failed to save settings to {SETTINGS_FILE}: {e}")
            QMessageBox.critical(
                self, "Error", f"Failed to save settings: {e}")

    def open_settings(self):
        """
        Opens the settings dialog and saves changes to the external settings file.
        """
        dialog = SettingsDialog(self.settings.copy(),
                                self)  # Pass a copy to allow cancellation
        if dialog.exec_() == QDialog.Accepted:
            new_settings = dialog.get_settings()
            if new_settings != self.settings:
                # Preserve the runtime stats_date before overwriting settings
                current_stats_date = self.settings.get("stats_date")

                # Save the new settings to the JSON file
                self.save_settings(new_settings)

                # Restore the runtime-only stats_date setting
                self.settings["stats_date"] = current_stats_date

                # Apply changes
                self.update_stats()
                if self.is_camera_active:
                    self.stop_camera()
                    self.start_camera()
                self.statusBar().showMessage("Settings updated", 5000)

    def init_ui(self):
        self.setWindowTitle(APP_TITLE)
        self.showFullScreen()
        self.apply_theme()

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(GROUP_SPACING)
        main_layout.setContentsMargins(15, 15, 15, 15)

        # Menu Bar
        menubar = self.menuBar()
        if not menubar:
            menubar = QMenuBar()
            self.setMenuBar(menubar)

        # File Menu
        file_menu = menubar.addMenu("&File")
        exit_action = QAction("&Exit", self)
        exit_action.setStatusTip("Exit the application")
        exit_action.setShortcut("Esc")
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)

        # Tools Menu
        tools_menu = menubar.addMenu("&Tools")
        preview_qr_action = QAction("&Preview QR Card", self)
        preview_qr_action.setStatusTip(
            "Show a preview of the QR code card design")
        preview_qr_action.triggered.connect(self.show_qr_preview)
        tools_menu.addAction(preview_qr_action)

        # Help Menu
        help_menu = menubar.addMenu("&Help")
        about_action = QAction("&About", self)
        about_action.setStatusTip("Show information about the application")
        about_action.triggered.connect(self.show_about_dialog)
        help_menu.addAction(about_action)

        # Create Tabs
        self.tabs = QTabWidget()
        scanner_tab = self.create_scanner_tab()
        if scanner_tab:
            self.tabs.addTab(scanner_tab, "Scanner")

        dashboard_tab = self.create_dashboard_tab()
        if dashboard_tab:
            self.tabs.addTab(dashboard_tab, "Dashboard")

        analytics_tab = self.create_analytics_tab()
        if analytics_tab:
            self.tabs.addTab(analytics_tab, "Analytics")

        main_layout.addWidget(self.tabs, stretch=1)

        # Timer for updating the clock
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.update_current_time)
        self.timer.start(1000)

        self.clear_video_frame()

    def apply_theme(self):
        """Applies the global visual theme and fonts."""
        font_size = FIXED_FONT_SIZE

        # Apply QSS colors and padding
        self.setStyleSheet(f"""
            QMainWindow, QDialog, QWidget {{ 
                background-color: {THEME_BG_COLOR}; 
                color: {TEXT_COLOR}; 
                font-family: "{FIXED_FONT_FAMILY}";
                font-size: {font_size}pt;
            }}
            QPushButton {{ 
                background-color: {BUTTON_COLOR}; 
                color: {BUTTON_TEXT_COLOR}; 
                padding: 8px; 
                border-radius: 6px; 
                border: 1px solid #BDC3C7; 
            }}
            QPushButton:hover {{ background-color: {BUTTON_HOVER_COLOR}; }}
            QTableWidget {{ 
                background-color: {THEME_BG_COLOR}; 
                color: {TEXT_COLOR}; 
                border: 1px solid #BDC3C7; 
                border-radius: 5px; 
                gridline-color: #E0E0E0; 
            }}
            QTableWidget::item {{ padding: 5px; }}
            QListWidget {{ 
                border: 1px solid #BDC3C7; 
                border-radius: 5px; 
                padding: 5px; 
                background-color: #FFFFFF; 
            }}
            QLineEdit, QComboBox, QDateEdit {{ 
                padding: 8px; 
                border: 1px solid #BDC3C7; 
                border-radius: 6px; 
                background-color: #FFFFFF; 
            }}
            QLabel {{ padding: 3px; }}
            QTabWidget::pane {{
                border: 1px solid #BDC3C7;
                border-radius: 8px;
                background: white;
            }}
            QTabBar::tab {{
                background: #F5F6F5;
                border: 1px solid #BDC3C7;
                border-bottom: none;
                border-top-left-radius: 8px;
                border-top-right-radius: 8px;
                padding: 8px 16px;
                margin-right: 2px;
            }}
            QTabBar::tab:selected {{
                background: white;
                border-bottom: 2px solid #0984E3;
            }}
            QTabBar::tab:hover {{
                background: #E0E0E0;
            }}
            QGroupBox {{
                border: 1px solid #BDC3C7;
                border-radius: 8px;
                margin-top: 10px;
                background-color: {THEME_BG_COLOR};
            }}
            QGroupBox::title {{
                subcontrol-origin: margin;
                subcontrol-position: top center;
                padding: 0 10px;
            }}
        """)
        self.apply_fonts_recursively(self)

    def apply_fonts_recursively(self, widget):
        """Apply global font to all widgets recursively."""
        font = QFont(FIXED_FONT_FAMILY, FIXED_FONT_SIZE)
        widget.setFont(font)
        for child in widget.findChildren(QWidget):
            child.setFont(font)

    def create_footer(self, text="Developed by Sao Moni"):
        footer_label = QLabel(text)
        footer_label.setFont(QFont(UI_FONT_FAMILY, 8, QFont.Light))
        footer_label.setAlignment(Qt.AlignCenter)
        footer_label.setStyleSheet("""
            color: #999;
            font-style: italic;
            background: transparent;
            padding: 0px;
            margin: 0px;
        """)
        footer_widget = QWidget()
        layout = QHBoxLayout(footer_widget)
        layout.setContentsMargins(0, 2, 0, 2)
        layout.addWidget(footer_label, alignment=Qt.AlignCenter)
        footer_widget.setStyleSheet(
            "border-top: 1px solid #E0E0E0; background: transparent;")
        footer_widget.setFixedHeight(18)
        return footer_widget

    def create_scanner_tab(self):
        scanner_widget = QWidget()
        main_layout = QVBoxLayout(scanner_widget)
        main_layout.setSpacing(GROUP_SPACING)
        main_layout.setContentsMargins(15, 15, 15, 15)

        # Header
        header_widget = QWidget()
        header_widget.setFixedHeight(60)
        header_widget.setStyleSheet(f"""
            background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                stop:0 {BUTTON_COLOR}, stop:1 {BUTTON_HOVER_COLOR});
            border-radius: 12px;
        """)
        header_layout = QHBoxLayout(header_widget)
        header_label = QLabel("Smart Library Scanner")
        header_label.setFont(
            QFont(UI_FONT_FAMILY, int(FIXED_FONT_SIZE * 1.4), QFont.Bold))
        header_label.setStyleSheet(
            f"color: {BUTTON_TEXT_COLOR}; padding: 8px; background: transparent;")
        header_layout.addWidget(header_label, alignment=Qt.AlignCenter)
        main_layout.addWidget(header_widget)

        # Content
        content_widget = QWidget()
        content_layout = QHBoxLayout(content_widget)
        content_layout.setSpacing(30)

        # Fixed video size
        self.video_size = QSize(VIDEO_FIXED_SIZE[0], VIDEO_FIXED_SIZE[1])

        # Left Section: Video Feed
        left_section = QWidget()
        left_layout = QVBoxLayout(left_section)
        left_layout.setSpacing(10)

        video_section = QWidget()
        video_section.setStyleSheet(
            f"background-color: {VIDEO_BG_COLOR}; border-radius: 15px; padding: 10px; border: 1px solid #CCCCCC;")
        video_layout = QVBoxLayout(video_section)
        video_layout.setContentsMargins(5, 5, 5, 5)

        self.video_label = QLabel()
        self.video_label.setFixedSize(self.video_size)
        self.video_label.setScaledContents(True)
        self.video_label.setAlignment(Qt.AlignCenter)
        self.video_label.setStyleSheet(
            "background-color: #1E1E1E; border-radius: 10px;")
        video_layout.addWidget(self.video_label, alignment=Qt.AlignCenter)
        left_layout.addWidget(video_section)

        # Status Widget
        status_widget = QGroupBox("Scan Status")
        status_layout = QVBoxLayout(status_widget)
        status_layout.setSpacing(6)

        self.status_label = self.styled_label("Status: Ready", color="#4CAF50")
        self.user_info_label = self.styled_label("No user selected")
        self.status_indicator = self.styled_label(
            "Current Status: Unknown", color="#888")
        self.last_event_label = self.styled_label(
            "Last Event: None", color="#888")
        self.current_time_label = self.styled_label("", color="#888")

        for label in [self.status_label, self.user_info_label,
                      self.status_indicator, self.last_event_label,
                      self.current_time_label]:
            status_layout.addWidget(label)

        left_layout.addWidget(status_widget)
        left_layout.addStretch()
        content_layout.addWidget(left_section, stretch=3)

        # Right Section: Controls & Recent Activity
        right_section = QWidget()
        right_layout = QVBoxLayout(right_section)
        right_layout.setSpacing(15)
        right_layout.setAlignment(Qt.AlignTop)

        # Controls Group
        controls_group = QGroupBox("Controls")
        controls_layout = QVBoxLayout(controls_group)

        def styled_button(text, slot, checkable=False):
            button = QPushButton(text)
            button.setMinimumHeight(40)
            button.clicked.connect(slot)
            button.setFont(QFont(UI_FONT_FAMILY, FIXED_FONT_SIZE, QFont.Bold))
            if checkable:
                button.setCheckable(True)
            return button

        self.start_camera_button = styled_button(
            "Start Camera", self.toggle_camera, checkable=True)
        controls_layout.addWidget(self.start_camera_button)
        controls_layout.addWidget(styled_button(
            "User Management", self.open_user_management))

        self.manual_qr_input = QLineEdit()
        self.manual_qr_input.setPlaceholderText("Enter QR Code Manually...")
        self.manual_qr_input.returnPressed.connect(self.process_manual_qr)
        controls_layout.addWidget(self.manual_qr_input)
        right_layout.addWidget(controls_group)

        # Recent Activity Group
        activity_group = QGroupBox("Recent Activity")
        activity_layout = QVBoxLayout(activity_group)
        activity_layout.addWidget(self.recent_list)
        # Allow it to stretch
        right_layout.addWidget(activity_group, stretch=1)

        content_layout.addWidget(right_section, stretch=2)

        main_layout.addWidget(content_widget)
        main_layout.addStretch(1)
        main_layout.addWidget(self.create_footer())
        return scanner_widget

    def create_dashboard_tab(self):
        dashboard_widget = QWidget()
        main_layout = QGridLayout(dashboard_widget)
        main_layout.setSpacing(15)  # GROUP_SPACING
        main_layout.setContentsMargins(15, 15, 15, 15)

        # --- Style Definitions for a Flat Look ---
        # Style for cards: removes shadow, adds a simple border
        card_style = """
            border: 1px solid #E0E0E0;
            border-radius: 8px;
            background-color: white;
        """
        # Style for group boxes: replaces the default groove/ridge border with a flat line
        group_box_style = """
            QGroupBox {
                font-weight: bold;
                border: 1px solid #E0E0E0;
                border-radius: 8px;
                margin-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                subcontrol-position: top center;
                padding: 0 5px;
                background-color: #F5F5F5; /* Match parent background */
            }
        """
        # Style for buttons: removes default 3D beveling
        flat_button_style = """
            QPushButton {
                border: 1px solid #CCCCCC;
                border-radius: 4px;
                padding: 6px 12px;
                background-color: #0984E3;
            }
            QPushButton:hover {
                background-color: #74B9FF;
            }
            QPushButton:pressed {
                background-color: #D0D0D0;
                border-style: inset;
            }
        """

        # --- Header ---
        header = QLabel("Library Analytics Dashboard")
        header.setFont(QFont("Segoe UI", int(10 * 1.6),
                       QFont.Bold))  # Dummy values
        header.setAlignment(Qt.AlignCenter)
        header.setStyleSheet(f"""
            background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #4A90E2, stop:1 #50E3C2); /* Example colors */
            color: white;
            border-radius: 12px;
            padding: 10px;
        """)
        main_layout.addWidget(header, 0, 0, 1, 3)

        # --- Date Selector ---
        date_widget = QWidget()
        date_layout = QHBoxLayout(date_widget)
        date_layout.setContentsMargins(0, 5, 0, 5)
        # Assuming styled_label exists
        date_layout.addWidget(self.styled_label("<b>Select Date:</b>"))
        self.stats_date_edit = QDateEdit(QDate.currentDate())
        self.stats_date_edit.setCalendarPopup(True)
        self.stats_date_edit.dateChanged.connect(self.update_stats)
        date_layout.addWidget(self.stats_date_edit)

        refresh_button = QPushButton("Refresh Data")
        refresh_button.setStyleSheet(flat_button_style)  # Apply flat style
        refresh_button.clicked.connect(self.update_stats)
        date_layout.addWidget(refresh_button)
        date_layout.addStretch()
        main_layout.addWidget(date_widget, 1, 0, 1, 3)

        # --- Key Metrics Section (Cards) ---
        self.total_users_card = DashboardCard("Total Users")
        self.total_users_label = self.styled_label(
            "0", font_size=28, bold=True)
        self.total_users_label.setAlignment(Qt.AlignCenter)
        self.total_users_card.addWidget(self.total_users_label)

        self.checkins_card = DashboardCard("Today's Check-ins")
        self.checkins_progress = ProgressRing()  # Assuming this is a custom widget
        self.checkins_card.addWidget(self.checkins_progress)

        self.female_card = DashboardCard("Female Check-ins")
        self.female_label = self.styled_label("0", font_size=28, bold=True)
        self.female_label.setAlignment(Qt.AlignCenter)
        self.female_card.addWidget(self.female_label)

        # Apply flat styling to all cards
        cards = [self.total_users_card, self.checkins_card, self.female_card]
        for i, card in enumerate(cards):
            card.setStyleSheet(card_style)
            # Failsafe: Explicitly remove any graphics effect like QGraphicsDropShadowEffect
            card.setGraphicsEffect(None)
            main_layout.addWidget(card, 2, i)

        # --- Grade Check-ins Table ---
        self.grade_card = DashboardCard("Grade Check-ins")
        self.grade_card.setStyleSheet(card_style)  # Apply flat style
        self.grade_card.setGraphicsEffect(None)   # Failsafe

        self.grade_table = QTableWidget()
        self.grade_table.setColumnCount(3)
        self.grade_table.setHorizontalHeaderLabels(
            ['Grade', 'Total', 'Female'])
        self.grade_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.grade_table.verticalHeader().setVisible(False)
        self.grade_table.setEditTriggers(QTableWidget.NoEditTriggers)
        # No border needed inside card
        self.grade_table.setStyleSheet("border: none; background: white;")
        self.grade_card.addWidget(self.grade_table)
        main_layout.addWidget(self.grade_card, 3, 0, 2,
                              3)  # Span 2 rows, 3 cols

        # --- Action Buttons Grouped ---
        actions_widget = QWidget()
        # Apply the flat button style to all QPushButton children of this widget
        actions_widget.setStyleSheet(flat_button_style)
        actions_layout = QHBoxLayout(actions_widget)
        actions_layout.setSpacing(20)

        # User & Data Management Group
        user_group = QGroupBox("User & Data Management")
        user_group.setStyleSheet(group_box_style)  # Apply flat style
        user_layout = QVBoxLayout(user_group)
        user_layout.addWidget(QPushButton(
            "Import Users", clicked=self.import_users_from_excel))
        user_layout.addWidget(QPushButton(
            "Download Template", clicked=self.download_user_template))
        user_layout.addWidget(QPushButton(
            "View Detailed Logs", clicked=self.view_logs))
        user_layout.addWidget(QPushButton(
            "View Top Visitors", clicked=self.open_top_users_dialog))
        actions_layout.addWidget(user_group)

        # Reports & System Group
        system_group = QGroupBox("Reports & System")
        system_group.setStyleSheet(group_box_style)  # Apply flat style
        system_layout = QVBoxLayout(system_group)
        system_layout.addWidget(QPushButton(
            "Generate Summary Report", clicked=self.generate_summary_report))
        system_layout.addWidget(QPushButton(
            "Export Class Data", clicked=self.export_class_data_dialog))
        system_layout.addWidget(QPushButton(
            "Backup Data", clicked=self.backup_data))
        system_layout.addWidget(QPushButton(
            "Settings", clicked=self.open_settings))
        actions_layout.addWidget(system_group)

        main_layout.addWidget(actions_widget, 5, 0, 1, 3)
        main_layout.setRowStretch(4, 1)  # Add stretch to push buttons down

        # Assuming create_footer exists
        main_layout.addWidget(self.create_footer(), 6, 0, 1, 3)
        self.update_stats()
        return dashboard_widget

    def create_analytics_tab(self):
        analytics_widget = QWidget()
        main_layout = QVBoxLayout(analytics_widget)
        main_layout.setSpacing(GROUP_SPACING)
        main_layout.setContentsMargins(15, 15, 15, 15)

        # Header
        header = QLabel("Advanced Analytics")
        header.setFont(QFont(UI_FONT_FAMILY, int(
            FIXED_FONT_SIZE * 1.6), QFont.Bold))
        header.setAlignment(Qt.AlignCenter)
        header.setStyleSheet(f"""
            background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                stop:0 {BUTTON_COLOR}, stop:1 {BUTTON_HOVER_COLOR});
            color: {BUTTON_TEXT_COLOR};
            border-radius: 12px;
            padding: 10px;
            margin-bottom: 10px;
        """)
        main_layout.addWidget(header)

        # Year Selector
        year_layout = QHBoxLayout()
        year_layout.addWidget(self.styled_label(
            "<b>Select Year for Trends:</b>"))
        self.trend_year_combo = QComboBox()
        current_year = datetime.datetime.now().year
        years = [str(y) for y in range(current_year - 5, current_year + 1)]
        self.trend_year_combo.addItems(years)
        self.trend_year_combo.setCurrentText(str(current_year))
        self.trend_year_combo.currentTextChanged.connect(
            self.update_analytics_charts)
        year_layout.addWidget(self.trend_year_combo)
        year_layout.addStretch()
        main_layout.addLayout(year_layout)

        # Charts Section
        charts_splitter = QSplitter(Qt.Horizontal)
        charts_splitter.setHandleWidth(10)
        charts_splitter.setStyleSheet(
            "QSplitter::handle { background-color: #BDC3C7; }")

        self.monthly_chart_view = QChartView()
        self.monthly_chart_view.setRenderHint(QPainter.Antialiasing)
        charts_splitter.addWidget(self.monthly_chart_view)

        self.grade_pie_chart_view = QChartView()
        self.grade_pie_chart_view.setRenderHint(QPainter.Antialiasing)
        charts_splitter.addWidget(self.grade_pie_chart_view)

        charts_splitter.setStretchFactor(0, 1)
        charts_splitter.setStretchFactor(1, 1)
        main_layout.addWidget(charts_splitter, stretch=1)

        # Top Users List
        top_users_group = QGroupBox("Top 10 Most Active Users (All Time)")
        top_users_layout = QVBoxLayout(top_users_group)
        self.top_users_list = QListWidget()
        top_users_layout.addWidget(self.top_users_list)
        main_layout.addWidget(top_users_group, stretch=1)

        self.update_analytics_charts()
        self.update_top_users()

        main_layout.addWidget(self.create_footer())
        return analytics_widget

    def update_analytics_charts(self):
        # Update Monthly Trend Chart
        selected_year = int(self.trend_year_combo.currentText())
        monthly_data = self.log_manager.get_monthly_trend_data(selected_year)

        series = QBarSeries()
        bar_set = QBarSet("Check-ins")
        for count in monthly_data:
            bar_set.append(count)
        series.append(bar_set)

        chart = QChart()
        chart.addSeries(series)
        chart.setTitle(f"Monthly Check-in Trend - {selected_year}")
        chart.setAnimationOptions(QChart.SeriesAnimations)

        axis_x = QBarCategoryAxis()
        months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                  "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        axis_x.append(months)
        chart.addAxis(axis_x, Qt.AlignBottom)
        series.attachAxis(axis_x)

        axis_y = QValueAxis()
        axis_y.setLabelFormat("%i")
        axis_y.setTitleText("Number of Check-ins")
        chart.addAxis(axis_y, Qt.AlignLeft)
        series.attachAxis(axis_y)

        chart.legend().setVisible(False)
        chart.setBackgroundRoundness(10)
        self.monthly_chart_view.setChart(chart)

        # Update Grade Distribution Chart for Selected Date
        selected_date = self.stats_date_edit.date().toPyDate()
        grade_status = self.log_manager.get_grade_level_current_status(
            selected_date)

        pie_series = QPieSeries()
        total_pie_checkins = 0
        for grade in range(7, 13):
            count = grade_status.get(grade, {}).get('total_checked_in', 0)
            if count > 0:
                total_pie_checkins += count
                pie_series.append(f"Grade {grade}", count)

        # Customize slices
        for slice in pie_series.slices():
            slice.setLabel(f"{slice.label()} ({slice.value():.0f})")

        pie_chart = QChart()
        pie_chart.addSeries(pie_series)
        pie_chart.setTitle(
            f"Grade Distribution - {selected_date.strftime('%Y-%m-%d')}")
        pie_chart.setAnimationOptions(QChart.SeriesAnimations)
        pie_chart.legend().setAlignment(Qt.AlignRight)
        pie_chart.setBackgroundRoundness(10)

        self.grade_pie_chart_view.setChart(pie_chart)

    def clear_video_frame(self):
        pixmap = QPixmap(self.video_size)
        pixmap.fill(QColor(VIDEO_BG_COLOR))
        painter = QPainter(pixmap)
        painter.setFont(QFont(UI_FONT_FAMILY, 14, QFont.Bold))
        painter.setPen(QColor(TEXT_COLOR))
        painter.drawText(pixmap.rect(), Qt.AlignCenter,
                         "Camera Inactive\nPress 'Start Camera' to Scan")
        painter.end()
        self.video_label.setPixmap(pixmap)

    def update_video_frame(self, pixmap):
        scaled_pixmap = pixmap.scaled(
            self.video_size, Qt.KeepAspectRatio, Qt.SmoothTransformation)
        final_pixmap = QPixmap(self.video_size)
        final_pixmap.fill(QColor("#1E1E1E"))
        painter = QPainter(final_pixmap)
        offset_x = (self.video_size.width() - scaled_pixmap.width()) // 2
        offset_y = (self.video_size.height() - scaled_pixmap.height()) // 2
        painter.drawPixmap(offset_x, offset_y, scaled_pixmap)
        painter.end()
        self.video_label.setPixmap(final_pixmap)

 # In class QRCodeApp
    def check_sound_files(self):
        sound_files = [INVAILD_SOUND, CHECK_IN_SOUND, CHECK_OUT_SOUND]
        missing_files = [f for f in sound_files if not os.path.exists(f)]

        # This is the corrected logic:
        # Only force-disable the sound if it was enabled in the settings AND files are missing.
        if missing_files and self.settings.get("sound_enabled", False):
            logging.warning(
                f"Sound files missing: {', '.join(missing_files)}. Sound has been disabled for this session.")
            self.settings["sound_enabled"] = False
            self.statusBar().showMessage(
                f"Warning: Sound files are missing, so sound has been auto-disabled.", 10000)

    def check_camera_availability(self):
        self.available_cameras = []
        for i in range(10):
            cap = cv2.VideoCapture(i)
            if cap.isOpened():
                self.available_cameras.append(i)
                cap.release()
        if not self.available_cameras:
            logging.warning("No webcams detected")
            self.statusBar().showMessage("No webcams detected", 5000)
        else:
            logging.info(
                f"Detected cameras at indices: {self.available_cameras}")
            if self.settings["camera_index"] not in self.available_cameras:
                self.settings["camera_index"] = self.available_cameras[0]
                logging.info(
                    f"Defaulted to camera index {self.settings['camera_index']}")

    def toggle_camera(self):
        if self.start_camera_button.isChecked():
            self.start_camera()
        else:
            self.stop_camera()

    def handle_scanned_code(self, decoded_text):
        if self.scan_cooldown or not decoded_text.strip():
            return
        if decoded_text == self.last_scanned_code:
            return
        self.decoded_data = decoded_text
        self.last_scanned_code = decoded_text

    def process_scanning(self):
        if self.scan_cooldown or not getattr(self, 'decoded_data', None):
            return

        self.scan_cooldown = True
        self.scan_timer.stop()
        self.status_label.setText("Status: Processing...")

        decoded_text = self.decoded_data
        self.decoded_data = None

        user_id_match = re.search(r"user_id:(\d+)", decoded_text)
        if not user_id_match:
            self.statusBar().showMessage("Invalid QR code format", 5000)
            self.play_sound(INVAILD_SOUND)
            QTimer.singleShot(1000, self.restart_scanning)
            return

        user_id = int(user_id_match.group(1))
        user = self.user_manager.get_user_by_id(user_id)
        if not user:
            self.statusBar().showMessage(f"User ID {user_id} not found", 5000)
            self.play_sound(INVAILD_SOUND)
            QTimer.singleShot(1000, self.restart_scanning)
            return

        last_event = self.log_manager.get_last_event_for_user(user_id)
        event_type = "Check-out" if last_event == "Check-in" else "Check-in"

        timestamp = self.log_manager.log_event(user_id, event_type)
        if timestamp:
            self.current_user_id = user_id
            self.update_user_info(user)
            self.last_event_label.setText(
                f"Last Event: {event_type} at {timestamp}")
            sound_file = CHECK_IN_SOUND if event_type == "Check-in" else CHECK_OUT_SOUND
            self.play_sound(sound_file)
            self.statusBar().showMessage(
                f"{event_type} recorded for {user['first_name']} {user['last_name']}", 5000
            )
            self.update_stats()
        else:
            self.statusBar().showMessage("Failed to log event.", 5000)
            self.play_sound(INVAILD_SOUND)

        QTimer.singleShot(1500, self.restart_scanning)

    def restart_scanning(self):
        self.scan_cooldown = False
        self.last_scanned_code = None
        self.decoded_data = None
        self.status_label.setText("Status: Ready")
        if self.is_camera_active:
            self.scan_timer.start(self.scan_interval)

    def play_sound(self, sound_file):
        if self.settings["sound_enabled"] and os.path.exists(sound_file):
            try:
                self.media_player.setMedia(QMediaContent(
                    QUrl.fromLocalFile(os.path.abspath(sound_file))))
                self.media_player.play()
            except Exception as e:
                logging.error(f"Failed to play sound {sound_file}: {e}")

    def update_user_info(self, user_data):
        try:
            if user_data:
                self.user_info_label.setText(
                    f"User: {user_data['first_name']} {user_data['last_name']} "
                    f"(ID: {user_data['id']}, Class: {user_data.get('class', 'N/A')})"
                )
                last_event = self.log_manager.get_last_event_for_user(
                    user_data['id'])
                if last_event == "Check-in":
                    status_text, style = "Checked In", "background-color: #4CAF50; color: white; padding: 6px; border-radius: 5px;"
                elif last_event == "Check-out":
                    status_text, style = "Checked Out", "background-color: #F44336; color: white; padding: 6px; border-radius: 5px;"
                else:
                    status_text, style = "Unknown", "background-color: #E0E0E0; color: #757575; padding: 6px; border-radius: 5px;"
                self.status_indicator.setText(f"Status: {status_text}")
                self.status_indicator.setStyleSheet(style)
            else:
                self.user_info_label.setText("No user selected")
                self.status_indicator.setText("Status: Unknown")
                self.status_indicator.setStyleSheet(
                    "background-color: #E0E0E0; color: #757575; padding: 6px; border-radius: 5px;")
        except Exception as e:
            logging.error(f"Failed to update user info: {e}")
            self.statusBar().showMessage("Error updating user info", 5000)

    def update_current_time(self):
        self.current_time_label.setText(
            f"Time: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    def update_stats_date(self, qdate):
        self.settings["stats_date"] = qdate.toPyDate()
        self.update_stats()

    def update_stats(self):
        try:
            users = self.user_manager.get_all_users()
            total_users = len(users)
            selected_date = self.stats_date_edit.date().toPyDate()
            self.settings["stats_date"] = selected_date

            # This now gets unique counts for both total and female check-ins
            unique_check_ins, unique_female_check_ins = self.log_manager.get_daily_stats(
                selected_date)

            grade_status = self.log_manager.get_grade_level_current_status(
                selected_date)

            self.total_users_label.setText(f"{total_users}")

            # Update the progress ring with the unique user count
            daily_percentage = min(
                100, (unique_check_ins / 500) * 100) if 500 > 0 else 0
            self.checkins_progress.setValue(int(daily_percentage))
            self.checkins_progress.setText(f"{unique_check_ins}")

            # Update the female check-ins label with the unique female count
            self.female_label.setText(f"{unique_female_check_ins}")

            self.grade_table.setRowCount(6)
            for i, grade in enumerate(range(7, 13)):
                total = grade_status.get(grade, {}).get('total_checked_in', 0)
                female = grade_status.get(grade, {}).get(
                    'female_checked_in', 0)
                self.grade_table.setItem(
                    i, 0, QTableWidgetItem(f"Grade {grade}"))
                self.grade_table.setItem(i, 1, QTableWidgetItem(str(total)))
                self.grade_table.setItem(i, 2, QTableWidgetItem(str(female)))

            self.update_recent_activity()
            self.update_analytics_charts()
            self.statusBar().showMessage("Stats refreshed.", 3000)
        except Exception as e:
            logging.error(f"Failed to update stats: {e}")
            self.statusBar().showMessage("Failed to refresh stats.", 5000)

    def update_recent_activity(self):
        self.recent_list.clear()
        recent_events = self.log_manager.get_recent_events()
        for user_id, event, timestamp in recent_events:
            user = self.user_manager.get_user_by_id(user_id)
            if user:
                item = QListWidgetItem(
                    f"{event} for {user['first_name']} {user['last_name']}\n@ {timestamp}")
                color = QColor(
                    BUTTON_COLOR) if event == "Check-in" else QColor("#E74C3C")
                item.setForeground(color)
                self.recent_list.addItem(item)

    def update_top_users(self):
        self.top_users_list.clear()
        top_checkins = self.log_manager.get_top_users_by_checkins(limit=10)
        if not top_checkins:
            self.top_users_list.addItem("No data available.")
            return

        for rank, (user_id, count) in enumerate(top_checkins, 1):
            user = self.user_manager.get_user_by_id(user_id)
            if user:
                self.top_users_list.addItem(
                    f"#{rank}: {user['first_name']} {user['last_name']} (Total: {count})")
            else:
                self.top_users_list.addItem(
                    f"#{rank}: User ID {user_id} (Total: {count}) - Data missing")

    def process_manual_qr(self):
        qr_text = self.manual_qr_input.text().strip()
        if qr_text:
            self.handle_scanned_code(qr_text)
            self.status_label.setText("Status: Processing Manual Input")
            QTimer.singleShot(
                1000, lambda: self.status_label.setText("Status: Ready"))
            self.manual_qr_input.clear()

    def show_about_dialog(self):
        AboutDialog(self).exec_()

    def show_qr_preview(self):
        dummy_user = {
            "id": 11152,
            "first_name": "សៅ",
            "last_name": "មុន្នី",
            "class": "12B7",
            "gender": "ប្រុស",
            "role": "Student",
            "study_year": self.settings.get("default_study_year", "2025-2026")
        }
        dialog = QRCodeDialog(dummy_user, self.settings, self)
        dialog.exec_()

    def open_top_users_dialog(self):
        dialog = TopUsersDialog(
            self.user_manager, self.log_manager, self.settings, self)
        if dialog.exec_() == QDialog.Accepted:
            self.update_top_users()

    def start_camera(self):
        if self.is_camera_active:
            return
        if not self.available_cameras:
            self.statusBar().showMessage("No cameras available", 5000)
            self.start_camera_button.setChecked(False)
            return

        self.is_camera_active = True
        self.start_camera_button.setText("Pause Camera")
        self.statusBar().showMessage("Camera started", 5000)

        self.camera_thread = CameraThread(self, self.settings["camera_index"])
        self.camera_thread.finished_frame.connect(self.update_video_frame)
        self.camera_thread.scanned_code.connect(self.handle_scanned_code)
        self.camera_thread.start()
        self.scan_timer.start(self.scan_interval)

    def stop_camera(self):
        if not self.is_camera_active:
            return

        self.is_camera_active = False
        self.start_camera_button.setText("Start Camera")
        self.statusBar().showMessage("Camera stopped", 3000)

        if self.camera_thread:
            self.camera_thread.stop()
            self.camera_thread.wait()
            self.camera_thread = None

        self.scan_timer.stop()
        self.clear_video_frame()

    def open_user_management(self):
        UserManagementWindow(self.user_manager, self.settings, self).exec_()
        self.update_stats()

    def import_users_from_excel(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Import Users", "", "Excel Files (*.xlsx *.xls)")
        if file_path:
            try:
                df = pd.read_excel(file_path)
                required_columns = {'id', 'first_name',
                                    'last_name', 'class', 'gender', 'role'}
                if not required_columns.issubset(df.columns):
                    raise ValueError(
                        "Excel file missing required columns: id, first_name, last_name, class, gender, role")

                users = self.user_manager.load_users()
                existing_ids = {u['id'] for u in users}
                skipped_users = []
                new_users = 0

                for _, row in df.iterrows():
                    user_data = row.to_dict()
                    user_id = int(user_data['id']) if pd.notna(
                        user_data['id']) else None
                    if user_id in existing_ids:
                        skipped_users.append(f"ID {user_id}")
                        continue

                    user_data['id'] = user_id
                    user_data['study_year'] = self.settings.get(
                        "default_study_year", DEFAULT_STUDY_YEAR)
                    self.user_manager.add_user(user_data)
                    new_users += 1

                QMessageBox.information(
                    self, "Import Complete", f"{new_users} user(s) imported successfully.")
                self.update_stats()

                if skipped_users:
                    SkippedUsersDialog(skipped_users, self).exec_()

            except Exception as e:
                logging.error(f"Import failed: {e}")
                QMessageBox.critical(
                    self, "Import Failed", f"An error occurred during import: {str(e)}")

    def download_user_template(self):
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save Template", "user_data_template.xlsx", "Excel Files (*.xlsx)")
        if file_path:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(['id', 'first_name', 'last_name',
                         'class', 'gender', 'role', 'study_year'])
            sheet.append([1, 'សៅ', 'មុន្នី', '7A', 'ប្រុស', 'Student', self.settings.get(
                "default_study_year", DEFAULT_STUDY_YEAR)])
            workbook.save(file_path)
            self.statusBar().showMessage(
                f"Template saved to: {file_path}", 5000)

    def view_logs(self):
        LogViewerDialog(self.log_manager, self.user_manager,
                        self.settings, self).exec_()
        self.update_stats()

    def generate_summary_report(self):
        from datetime import date, datetime as dt
        from reportlab.platypus import Image

        start_date = self.settings["stats_date"]
        end_date = self.settings["stats_date"]

        filename = f"summary_report_{start_date.strftime('%Y%m%d')}.pdf"
        filepath, _ = QFileDialog.getSaveFileName(self, "Save Summary Report", os.path.join(
            self.settings.get("qr_save_dir", "."), filename), "PDF Files (*.pdf)")
        if not filepath:
            return

        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        doc = SimpleDocTemplate(filepath, pagesize=letter, topMargin=50,
                                bottomMargin=50, leftMargin=50, rightMargin=50)
        elements, styles = [], getSampleStyleSheet()

        title_style = styles['h1']
        title_style.alignment = 1
        subtitle_style = styles['h2']
        subtitle_style.alignment = 1
        normal_style = styles['Normal']

        # Header
        elements.append(
            Paragraph("Samdech Ov Samdech Mae High School", title_style))
        elements.append(Paragraph(
            "Smart Library System - Grade-Level Attendance Report", subtitle_style))
        elements.append(
            Paragraph(f"Date: {start_date.strftime('%Y-%m-%d')}", normal_style))
        elements.append(Spacer(1, 20))

        # Table
        grade_status = self.log_manager.get_grade_level_current_status(
            start_date)
        table_data = [['Grade', 'Total Checked In', 'Female Checked In']]
        for grade in range(7, 13):
            table_data.append([
                f"Grade {grade}",
                str(grade_status.get(grade, {}).get('total_checked_in', 0)),
                str(grade_status.get(grade, {}).get('female_checked_in', 0))
            ])

        grade_table = Table(table_data, colWidths=[150, 150, 150])
        grade_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(grade_table)
        elements.append(Spacer(1, 30))

        # Footer
        elements.append(Paragraph(
            f"Generated on: {dt.now().strftime('%Y-%m-%d %H:%M:%S')}", normal_style))

        try:
            doc.build(elements)
            QMessageBox.information(
                self, "Report Generated", f"Summary report exported to:\n{filepath}")
        except Exception as e:
            QMessageBox.critical(
                self, "Error", f"Failed to generate report: {str(e)}")

    def export_class_data_dialog(self):
        if not self.user_manager.get_classes():
            QMessageBox.information(
                self, "No Classes", "No classes available to export")
            return
        ExportClassDialog(self.user_manager, self.log_manager,
                          self.settings, self).exec_()
        self.update_stats()

    def backup_data(self):
        backup_dir = os.path.join(self.settings.get(
            "qr_save_dir", "."), f"backup_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}")
        os.makedirs(backup_dir, exist_ok=True)
        try:
            if os.path.exists(USER_DATA_FILE):
                shutil.copy(USER_DATA_FILE, backup_dir)
            if os.path.exists(LOG_DATABASE_FILE):
                shutil.copy(LOG_DATABASE_FILE, backup_dir)
            QMessageBox.information(
                self, "Backup Complete", f"Data backed up to:\n{backup_dir}")
        except Exception as e:
            logging.error(f"Failed to backup data: {e}")
            QMessageBox.critical(
                self, "Error", f"Failed to backup data: {str(e)}")

    def closeEvent(self, event):
        if self.is_camera_active:
            self.stop_camera()
        if self.log_manager:
            try:
                self.log_manager.close()
            except Exception as e:
                logging.error(f"Failed to close log manager: {e}")
        event.accept()

# Camera Thread


class CameraThread(QThread):
    finished_frame = pyqtSignal(QPixmap)
    scanned_code = pyqtSignal(str)

    def __init__(self, parent, camera_index):
        super().__init__(parent)
        self.camera_index = camera_index
        self.running = True
        self.qr_decoder = cv2.QRCodeDetector()

    def run(self):
        cap = cv2.VideoCapture(self.camera_index)
        cap.set(cv2.CAP_PROP_FPS, CAMERA_FPS)
        if not cap.isOpened():
            logging.error(
                f"Failed to open camera at index {self.camera_index}")
            self.running = False
            return

        while self.running and cap.isOpened():
            ret, frame = cap.read()
            if not ret or frame is None or frame.size == 0:
                time.sleep(0.1)
                continue

            frame = cv2.flip(frame, 1)
            rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            h, w, ch = rgb_frame.shape
            q_image = QImage(rgb_frame.data, w, h, ch *
                             w, QImage.Format_RGB888)
            pixmap = QPixmap.fromImage(q_image)
            self.finished_frame.emit(pixmap)

            try:
                decoded_data, _, _ = self.qr_decoder.detectAndDecode(frame)
                if decoded_data:
                    self.scanned_code.emit(decoded_data)
            except cv2.error:
                pass  # Ignore detection errors

            time.sleep(1 / CAMERA_FPS)
        cap.release()

    def stop(self):
        self.running = False
        self.wait()


class SettingsDialog(QDialog):
    def __init__(self, current_settings, parent=None):
        super().__init__(parent)
        self.settings = current_settings
        self.setWindowTitle("Settings")
        self.setMinimumSize(400, 300)
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(GROUP_SPACING)
        layout.setContentsMargins(
            DIALOG_PADDING, DIALOG_PADDING, DIALOG_PADDING, DIALOG_PADDING)

        # Sound feedback
        self.sound_checkbox = QCheckBox("Enable Sound Feedback")
        self.sound_checkbox.setChecked(
            self.settings.get("sound_enabled", True))
        layout.addWidget(self.sound_checkbox)

        # Camera selection
        layout.addWidget(QLabel("Select Camera:"))
        self.camera_combo = QComboBox()
        self.detect_cameras()
        self.camera_combo.setCurrentIndex(self.settings.get("camera_index", 0))
        layout.addWidget(self.camera_combo)

        # Study year
        layout.addWidget(QLabel("Default Study Year:"))
        self.study_year_input = QLineEdit()
        self.study_year_input.setText(
            self.settings.get("default_study_year", "2025-2026"))
        layout.addWidget(self.study_year_input)

        # QR save directory
        layout.addWidget(QLabel("QR Code & Report Save Directory:"))
        qr_dir_layout = QHBoxLayout()
        self.qr_dir_input = QLineEdit()
        self.qr_dir_input.setText(self.settings.get(
            "qr_save_dir", "library_data/qr_codes"))
        self.qr_dir_browse_btn = QPushButton("Browse")
        self.qr_dir_browse_btn.clicked.connect(self.browse_qr_dir)
        qr_dir_layout.addWidget(self.qr_dir_input)
        qr_dir_layout.addWidget(self.qr_dir_browse_btn)
        layout.addLayout(qr_dir_layout)

        layout.addStretch()

        # Save button
        save_button = QPushButton("Save")
        save_button.clicked.connect(self.accept)
        layout.addWidget(save_button, alignment=Qt.AlignCenter)

    def detect_cameras(self):
        self.camera_combo.clear()
        for i in range(10):
            cap = cv2.VideoCapture(i)
            if cap.isOpened():
                self.camera_combo.addItem(f"Camera {i}")
                cap.release()
            else:
                break
        if self.camera_combo.count() == 0:
            self.camera_combo.addItem("No cameras detected")
            self.camera_combo.setEnabled(False)

    def browse_qr_dir(self):
        folder = QFileDialog.getExistingDirectory(
            self, "Select Save Directory")
        if folder:
            self.qr_dir_input.setText(folder)

    # In class SettingsDialog
    def get_settings(self):
        """Returns a dictionary of ONLY the settings managed by this dialog."""
        return {
            "sound_enabled": self.sound_checkbox.isChecked(),
            "camera_index": self.camera_combo.currentIndex() if self.camera_combo.isEnabled() else 0,
            "default_study_year": self.study_year_input.text().strip(),
            "qr_save_dir": self.qr_dir_input.text().strip()
        }


class AboutDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"About {APP_TITLE}")
        self.setFixedSize(500, 480)  # Slightly taller for better spacing
        self.setWindowFlags(self.windowFlags() & ~
                            Qt.WindowContextHelpButtonHint)
        self.init_ui()

    def init_ui(self):
        # Main layout with a gradient background
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        self.setStyleSheet(f"""
            AboutDialog {{
                background-color: qlineargradient(
                    x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 {THEME_BG_COLOR}, stop: 1 #FFFFFF
                );
            }}
        """)

        # --- Central Card Widget for content ---
        card_widget = QFrame(self)
        card_widget.setStyleSheet("""
            QFrame {
                background-color: rgba(255, 255, 255, 0.9);
                border-radius: 15px;
            }
        """)
        card_layout = QVBoxLayout(card_widget)
        card_layout.setAlignment(Qt.AlignCenter)
        card_layout.setSpacing(15)
        card_layout.setContentsMargins(25, 25, 25, 25)

        # --- Title and Logo ---
        title_layout = QHBoxLayout()
        title_layout.setAlignment(Qt.AlignCenter)
        try:
            logo = QPixmap("asset/logo/logo.png")
            if not logo.isNull():
                logo_label = QLabel()
                logo_label.setPixmap(logo.scaled(
                    70, 70, Qt.KeepAspectRatio, Qt.SmoothTransformation))

                # Add shadow to the logo
                shadow = QGraphicsDropShadowEffect(self)
                shadow.setBlurRadius(15)
                shadow.setColor(QColor(0, 0, 0, 80))
                shadow.setOffset(2, 2)
                logo_label.setGraphicsEffect(shadow)

                title_layout.addWidget(logo_label)
        except Exception as e:
            logging.debug(f"Failed to load logo: {e}")

        title_info_layout = QVBoxLayout()
        app_name = QLabel(APP_TITLE)
        app_name.setFont(QFont(UI_FONT_FAMILY, 18, QFont.Bold))
        app_name.setStyleSheet(f"color: {BUTTON_COLOR};")

        version = QLabel("Version 2.2.0 - UI Refresh & Bug Fixes")
        version.setFont(QFont(UI_FONT_FAMILY, 10))
        version.setStyleSheet(f"color: {TEXT_COLOR};")

        title_info_layout.addWidget(app_name)
        title_info_layout.addWidget(version)
        title_layout.addLayout(title_info_layout)
        card_layout.addLayout(title_layout)

        # --- Separator ---
        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setFrameShadow(QFrame.Sunken)
        card_layout.addWidget(line)

        # --- Description with HTML & Icons ---
        about_text = QTextBrowser()
        about_text.setOpenExternalLinks(True)
        about_text.setStyleSheet("border: none; background: transparent;")

        # Get the current year dynamically
        current_year = datetime.date.today().year

        about_text.setHtml(f"""
            <p style="text-align: center; font-size: 11pt;">
                A data-driven attendance and analytics platform for the
                <b>Samdech Ov Samdech Mae High School Library</b>.
            </p>
            <h4 style="text-align: center; color: {BUTTON_COLOR};">Key Features</h4>
            <ul style="list-style-type: none; padding-left: 0; text-align: left; margin: auto; width: 80%;">
                <li> &nbsp; <b>Advanced Analytics:</b> Interactive charts for trends.</li>
                <li> &nbsp; <b>Real-time Dashboard:</b> At-a-glance metrics.</li>
                <li> &nbsp; <b>QR Code Scanning:</b> Fast check-in/check-out.</li>
                <li> &nbsp; <b>Comprehensive Reporting:</b> PDF and Excel exports.</li>
            </ul>
            <p style="text-align: center; color: #555; font-size: 9pt;">
                <br>Developed by <b>Sao Moni</b><br>
                Contact: <a href="mailto:keomony074@gmail.com" style="color: {BUTTON_COLOR}; text-decoration: none;">keomony074@gmail.com</a>
                <br>&copy; {current_year} All rights reserved.
            </p>
        """)
        card_layout.addWidget(about_text)

        # --- OK Button ---
        ok_button = QPushButton("OK")
        ok_button.setFixedWidth(120)
        ok_button.setFixedHeight(35)
        ok_button.clicked.connect(self.accept)
        card_layout.addWidget(ok_button, alignment=Qt.AlignCenter)

        main_layout.addWidget(card_widget)
# Top Users Dialog


class TopUsersDialog(QDialog):
    def __init__(self, user_manager, log_manager, settings, parent=None):
        super().__init__(parent)
        self.user_manager = user_manager
        self.log_manager = log_manager
        self.settings = settings
        self.setWindowTitle("Top 10 Most Frequent Visitors")
        self.setMinimumSize(500, 600)
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(GROUP_SPACING)
        layout.setContentsMargins(
            DIALOG_PADDING, DIALOG_PADDING, DIALOG_PADDING, DIALOG_PADDING)

        header = QLabel("Top 10 Most Frequent Visitors")
        header.setFont(QFont(UI_FONT_FAMILY, 14, QFont.Bold))
        header.setAlignment(Qt.AlignCenter)
        layout.addWidget(header)

        # Mode Selection
        mode_layout = QHBoxLayout()
        self.by_day_radio = QRadioButton("By Day")
        self.by_month_radio = QRadioButton("By Month")
        self.by_year_radio = QRadioButton("By Year")
        self.all_time_radio = QRadioButton("All Time")
        self.by_day_radio.setChecked(True)
        for radio in [self.by_day_radio, self.by_month_radio, self.by_year_radio, self.all_time_radio]:
            radio.toggled.connect(self.toggle_date_widgets)
            mode_layout.addWidget(radio)
        layout.addLayout(mode_layout)

        # Date Input Widgets
        self.day_input_widget = QDateEdit(QDate.currentDate())
        self.day_input_widget.setCalendarPopup(True)
        layout.addWidget(self.day_input_widget)

        self.month_input_widget = QWidget()
        month_layout = QHBoxLayout(self.month_input_widget)
        self.month_combo = QComboBox()
        self.month_combo.addItems(["January", "February", "March", "April", "May",
                                  "June", "July", "August", "September", "October", "November", "December"])
        self.month_combo.setCurrentIndex(QDate.currentDate().month() - 1)
        self.year_combo = QComboBox()
        current_year = QDate.currentDate().year()
        years = [str(y) for y in range(current_year - 5, current_year + 2)]
        self.year_combo.addItems(years)
        self.year_combo.setCurrentText(str(current_year))
        month_layout.addWidget(self.month_combo)
        month_layout.addWidget(self.year_combo)
        layout.addWidget(self.month_input_widget)

        self.year_input_widget = QComboBox()
        self.year_input_widget.addItems(years)
        self.year_input_widget.setCurrentText(str(current_year))
        layout.addWidget(self.year_input_widget)

        self.toggle_date_widgets()  # Set initial visibility

        # View Button
        view_button = QPushButton("View Top Visitors")
        view_button.clicked.connect(self.search_top_users)
        layout.addWidget(view_button)

        # Results List
        self.results_list = QListWidget()
        layout.addWidget(self.results_list)

        # Close Button
        close_button = QPushButton("Close")
        close_button.clicked.connect(self.accept)
        layout.addWidget(close_button, alignment=Qt.AlignCenter)

        self.search_top_users()

    def toggle_date_widgets(self):
        self.day_input_widget.setVisible(self.by_day_radio.isChecked())
        self.month_input_widget.setVisible(self.by_month_radio.isChecked())
        self.year_input_widget.setVisible(self.by_year_radio.isChecked())

    def search_top_users(self):
        self.results_list.clear()
        start_date, end_date = None, None

        if self.by_day_radio.isChecked():
            start_date = end_date = self.day_input_widget.date().toPyDate()
        elif self.by_month_radio.isChecked():
            month = self.month_combo.currentIndex() + 1
            year = int(self.year_combo.currentText())
            start_date = datetime.date(year, month, 1)
            last_day = (start_date.replace(month=month % 12 + 1, day=1) -
                        datetime.timedelta(days=1)).day if month < 12 else 31
            end_date = datetime.date(year, month, last_day)
        elif self.by_year_radio.isChecked():
            year = int(self.year_input_widget.currentText())
            start_date = datetime.date(year, 1, 1)
            end_date = datetime.date(year, 12, 31)

        top_checkins = self.log_manager.get_top_users_by_checkins(
            limit=10, start_date=start_date, end_date=end_date)
        if not top_checkins:
            self.results_list.addItem(
                "No frequent visitors found for this period.")
            return

        for rank, (user_id, count) in enumerate(top_checkins, 1):
            user = self.user_manager.get_user_by_id(user_id)
            if user:
                self.results_list.addItem(
                    f"#{rank}: {user['first_name']} {user['last_name']} (Check-ins: {count})")
            else:
                self.results_list.addItem(
                    f"#{rank}: User ID {user_id} (Check-ins: {count}) - User data missing")

# User Management Dialog
# User Management Dialog


class UserManagementWindow(QDialog):
    def __init__(self, user_manager, settings, parent=None):
        super().__init__(parent)
        self.user_manager = user_manager
        self.settings = settings
        self.setWindowTitle("User Management")
        self.setMinimumSize(900, 600)
        self.init_ui()

    def init_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(10)
        main_layout.setContentsMargins(10, 10, 10, 10)

        header = QLabel("User Management")
        header.setFont(QFont(UI_FONT_FAMILY, 14, QFont.Bold))
        header.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(header)

        # Search
        search_layout = QHBoxLayout()
        search_layout.addWidget(QLabel("Search:"))
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText(
            "Search by ID, First Name, or Last Name...")
        self.search_input.textChanged.connect(self.filter_users)
        search_layout.addWidget(self.search_input)
        main_layout.addLayout(search_layout)

        # User Table
        self.user_table = QTableWidget()
        self.user_table.setColumnCount(7)
        self.user_table.setHorizontalHeaderLabels(
            ['ID', 'First Name', 'Last Name', 'Class', 'Gender', 'Role', 'Study Year'])
        self.user_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.user_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.user_table.setSelectionMode(QTableWidget.SingleSelection)
        self.user_table.setEditTriggers(QTableWidget.NoEditTriggers)
        main_layout.addWidget(self.user_table)

        # Buttons
        button_layout = QHBoxLayout()
        buttons = {
            "Add User": self.add_user_dialog,
            "Edit User": self.edit_user_dialog,
            "Delete User": self.delete_user,
            "Generate QR Code": self.generate_qr_code,
            "Batch QR Codes": self.generate_batch_qr_codes,
            "Close": self.accept
        }
        for text, slot in buttons.items():
            button = QPushButton(text)
            button.clicked.connect(slot)
            button_layout.addWidget(button)
        main_layout.addLayout(button_layout)

        self.load_users()

    def filter_users(self):
        search_text = self.search_input.text().lower().strip()
        for row in range(self.user_table.rowCount()):
            id_item = self.user_table.item(row, 0).text()
            fname_item = self.user_table.item(row, 1).text().lower()
            lname_item = self.user_table.item(row, 2).text().lower()
            is_visible = (
                search_text == "" or search_text in id_item or search_text in fname_item or search_text in lname_item)
            self.user_table.setRowHidden(row, not is_visible)

    def load_users(self):
        users = self.user_manager.get_all_users()
        self.user_table.setRowCount(len(users))
        for row, user in enumerate(users):
            self.user_table.setItem(row, 0, QTableWidgetItem(str(user['id'])))
            self.user_table.setItem(
                row, 1, QTableWidgetItem(user.get('first_name', '')))
            self.user_table.setItem(
                row, 2, QTableWidgetItem(user.get('last_name', '')))
            self.user_table.setItem(
                row, 3, QTableWidgetItem(user.get('class', '')))
            self.user_table.setItem(
                row, 4, QTableWidgetItem(user.get('gender', '')))
            self.user_table.setItem(
                row, 5, QTableWidgetItem(user.get('role', '')))
            self.user_table.setItem(
                row, 6, QTableWidgetItem(user.get('study_year', '')))
        self.filter_users()

    def add_user_dialog(self):
        dialog = AddEditUserDialog(self.user_manager, self.settings, self)
        if dialog.exec_() == QDialog.Accepted:
            try:
                self.user_manager.add_user(dialog.get_user_data())
                self.load_users()
            except ValueError as e:
                QMessageBox.warning(self, "Error", str(e))

    def edit_user_dialog(self):
        selected_row = self.user_table.currentRow()
        if selected_row < 0:
            QMessageBox.warning(self, "Selection Error",
                                "Please select a user to edit.")
            return
        user_id = int(self.user_table.item(selected_row, 0).text())
        user_data = self.user_manager.get_user_by_id(user_id)

        dialog = AddEditUserDialog(
            self.user_manager, self.settings, self, user_data)
        if dialog.exec_() == QDialog.Accepted:
            updated_data = dialog.get_user_data()
            users = self.user_manager.load_users()
            for i, user in enumerate(users):
                if user['id'] == user_id:
                    users[i].update(updated_data)
                    break
            self.user_manager.save_users(users)
            self.load_users()

    def delete_user(self):
        selected_row = self.user_table.currentRow()
        if selected_row < 0:
            QMessageBox.warning(self, "Selection Error",
                                "Please select a user to delete.")
            return
        user_id = int(self.user_table.item(selected_row, 0).text())
        if QMessageBox.question(self, "Confirm Delete", f"Delete user ID {user_id}?", QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
            if self.user_manager.delete_user(user_id):
                self.load_users()
            else:
                QMessageBox.warning(self, "Error", "Failed to delete user.")

    def generate_qr_code(self):
        selected_row = self.user_table.currentRow()
        if selected_row < 0:
            QMessageBox.warning(self, "Selection Error",
                                "Please select a user.")
            return
        user_id = int(self.user_table.item(selected_row, 0).text())
        user_data = self.user_manager.get_user_by_id(user_id)
        if user_data:
            # Assuming create_qr_card_pixmap and QRCodeDialog are defined elsewhere
            dialog = QRCodeDialog(user_data, self.settings, self)
            dialog.exec_()

    def generate_batch_qr_codes(self):
        """Opens the batch QR code generation dialog."""
        # This is the corrected line that passes the settings dictionary properly.
        dialog = BatchQRCodeDialog(self.user_manager, self.settings, self)
        dialog.exec_()

# Add/Edit User Dialog


class AddEditUserDialog(QDialog):
    def __init__(self, user_manager, settings, parent=None, user_data=None):
        super().__init__(parent)
        self.user_manager = user_manager
        self.settings = settings
        self.user_data = user_data
        self.setWindowTitle("Edit User" if user_data else "Add User")
        self.setMinimumSize(400, 500)
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(GROUP_SPACING)
        layout.setContentsMargins(
            DIALOG_PADDING, DIALOG_PADDING, DIALOG_PADDING, DIALOG_PADDING)

        grid = QGridLayout()
        grid.setSpacing(10)

        fields = {
            "User ID:": ("id_input", "User ID (auto if blank)"),
            "First Name:": ("first_name_input", "គោត-នាម"),
            "Last Name:": ("last_name_input", "នាម"),
            "Class:": ("class_input", "e.g., 7A"),
            "Gender:": ("gender_combo", ["ប្រុស", "ស្រី"]),
            "Role:": ("role_combo", ["Student", "Teacher", "Staff"]),
            "Study Year:": ("study_year_input", "e.g., 2025-2026")
        }

        for i, (label_text, (widget_name, content)) in enumerate(fields.items()):
            label = QLabel(label_text)
            if isinstance(content, list):
                widget = QComboBox()
                widget.addItems(content)
            else:
                widget = QLineEdit()
                widget.setPlaceholderText(content)

            setattr(self, widget_name, widget)
            grid.addWidget(label, i, 0)
            grid.addWidget(widget, i, 1)

        layout.addLayout(grid)
        self.id_input.setReadOnly(bool(self.user_data))

        if self.user_data:
            self.id_input.setText(str(self.user_data.get('id', '')))
            self.first_name_input.setText(self.user_data.get('first_name', ''))
            self.last_name_input.setText(self.user_data.get('last_name', ''))
            self.class_input.setText(self.user_data.get('class', ''))
            self.gender_combo.setCurrentText(self.user_data.get('gender', ''))
            self.role_combo.setCurrentText(self.user_data.get('role', ''))
            self.study_year_input.setText(self.user_data.get(
                'study_year', self.settings.get('default_study_year', '')))

        button_layout = QHBoxLayout()
        save_button = QPushButton("Save")
        save_button.clicked.connect(self.accept)
        cancel_button = QPushButton("Cancel")
        cancel_button.clicked.connect(self.reject)
        button_layout.addWidget(save_button)
        button_layout.addWidget(cancel_button)
        layout.addLayout(button_layout)

    def get_user_data(self):
        try:
            user_id = self.id_input.text().strip()
            user_id = int(user_id) if user_id else None
        except ValueError:
            user_id = None

        return {
            'id': user_id,
            'first_name': self.first_name_input.text().strip(),
            'last_name': self.last_name_input.text().strip(),
            'class': self.class_input.text().strip(),
            'gender': self.gender_combo.currentText(),
            'role': self.role_combo.currentText(),
            'study_year': self.study_year_input.text().strip() or self.settings.get('default_study_year', '')
        }
# --- New, Independent Card Generation Function ---
# --- 1. Independent Card Generation Function (Keep your new design) ---


def create_qr_card_pixmap(user_data):
    """
    Generates a QR code and draws a complete, styled user card using the
    new horizontal design.

    Args:
        user_data (dict): A dictionary containing the user's information.

    Returns:
        QPixmap: The generated card image. Returns an error image on failure.
    """
    try:
        # Generate QR Code
        qr = qrcode.QRCode(
            version=1, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=8, border=2)
        qr.add_data(f"user_id:{user_data.get('id', 'unknown')}")
        qr.make(fit=True)
        qr_img = qr.make_image(fill_color="black", back_color="white")

        buffer = BytesIO()
        qr_img.save(buffer, format="PNG")
        buffer.seek(0)

        qr_pixmap = QPixmap()
        if not qr_pixmap.loadFromData(buffer.getvalue()):
            raise ValueError("Failed to load QR code image data from buffer")

        # Draw the Card using QPainter (Your new horizontal style)
        card_width, card_height = 600, 400
        card = QPixmap(card_width, card_height)
        card.fill(QColor("white"))
        painter = QPainter(card)
        if not painter.isActive():
            raise RuntimeError(
                "Failed to activate QPainter for card rendering.")
        try:
            painter.setRenderHint(QPainter.Antialiasing)
            painter.setRenderHint(QPainter.TextAntialiasing)

            # Blue border background
            painter.setBrush(QColor("#0984E3"))
            painter.setPen(Qt.NoPen)
            painter.drawRoundedRect(
                QRect(0, 0, card_width, card_height), 20, 20)

            # White main card area
            painter.setBrush(QColor("white"))
            painter.drawRoundedRect(
                QRect(10, 10, card_width - 20, card_height - 20), 15, 15)

            # School Title
            painter.setFont(QFont(FONT_FAMILY, 16, QFont.Bold))
            painter.setPen(QColor("#0984E3"))
            painter.drawText(QRect(0, 25, card_width, 40),
                             Qt.AlignCenter, "Samdech Ov Samdech Mae High School")

            # User Info Text
            painter.setFont(QFont(FONT_FAMILY, 11))
            painter.setPen(QColor("#2D3436"))
            info_lines = [
                f"ID: {user_data.get('id', 'N/A')}",
                f"Name: {user_data.get('first_name', '')} {user_data.get('last_name', '')}",
                f"Class: {user_data.get('class', 'N/A')}",
                f"Gender: {user_data.get('gender', 'N/A')}",
                f"Role: {user_data.get('role', 'N/A')}",
                f"Year: {user_data.get('study_year', 'N/A')}"
            ]
            y_pos = 110
            for line in info_lines:
                painter.drawText(40, y_pos, line)
                y_pos += 28

            # QR Code Image
            qr_size = 180
            qr_x = card_width - qr_size - 40
            qr_y = (card_height - qr_size) // 2
            painter.drawPixmap(qr_x, qr_y, qr_pixmap.scaled(
                qr_size, qr_size, Qt.KeepAspectRatio, Qt.SmoothTransformation))

            # Footer Text
            painter.setFont(QFont(FONT_FAMILY, 10, QFont.Bold))
            painter.setPen(QColor("#0984E3"))
            painter.drawText(QRect(0, card_height - 40, card_width, 30),
                             Qt.AlignCenter, "Smart Library System • Scan to Check In/Out")
        finally:
            if painter.isActive():
                painter.end()
        return card
    except Exception as e:
        logging.error(
            f"Critical error creating card for user {user_data.get('id', 'N/A')}: {e}")
        error_pixmap = QPixmap(600, 400)
        error_pixmap.fill(QColor("#FFF0F0"))
        p = QPainter(error_pixmap)
        p.setPen(QColor("red"))
        p.drawText(error_pixmap.rect(), Qt.AlignCenter,
                   f"Error generating card:\n{e}")
        p.end()
        return error_pixmap

# --- 2. QR Code Card Dialog (Improved version with Save and Print) ---


class QRCodeDialog(QDialog):
    """Dialog that displays a single user's QR card for preview, saving, and printing."""

    def __init__(self, user_data, settings, parent=None):
        super().__init__(parent)
        self.user_data = user_data
        self.settings = settings
        self.setWindowTitle("QR Code Card Preview")
        self.setFixedSize(680, 560)  # Adjusted for buttons
        self.card_pixmap = create_qr_card_pixmap(self.user_data)
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        # Card Preview
        card_label = QLabel()
        card_label.setAlignment(Qt.AlignCenter)
        if self.card_pixmap and not self.card_pixmap.isNull():
            card_label.setPixmap(self.card_pixmap.scaled(
                600, 400, Qt.KeepAspectRatio, Qt.SmoothTransformation))
        else:
            card_label.setText("Error: Could not render card.")
        layout.addWidget(card_label)

        # Buttons (Restoring functionality from your old code)
        button_layout = QHBoxLayout()
        save_button = QPushButton("Save Image")
        save_button.clicked.connect(self.save_card)
        print_button = QPushButton("Print Card")
        print_button.clicked.connect(self.print_card)
        close_button = QPushButton("Close")
        close_button.clicked.connect(self.accept)

        button_layout.addStretch()
        button_layout.addWidget(save_button)
        button_layout.addWidget(print_button)
        button_layout.addWidget(close_button)
        button_layout.addStretch()

        layout.addLayout(button_layout)

    def save_card(self):
        default_filename = f"qr_card_{self.user_data.get('first_name', '')}_{self.user_data.get('last_name', '')}_{self.user_data.get('id', 'N_A')}.png"
        default_path = os.path.join(self.settings.get(
            "qr_save_dir", os.getcwd()), default_filename)

        filename, _ = QFileDialog.getSaveFileName(
            self, "Save QR Code Card", default_path, "PNG Files (*.png)")
        if filename:
            try:
                os.makedirs(os.path.dirname(filename), exist_ok=True)
                if self.card_pixmap.save(filename, "PNG"):
                    QMessageBox.information(
                        self, "Success", f"Card saved to:\n{filename}")
                else:
                    raise IOError("Failed to save the image file.")
            except Exception as e:
                QMessageBox.critical(
                    self, "Error", f"Could not save the card:\n{e}")

    def print_card(self):
        printer = QPrinter(QPrinter.HighResolution)
        printer.setPageSize(QPrinter.A6)
        printer.setOrientation(QPrinter.Landscape)  # For the horizontal card

        dialog = QPrintDialog(printer, self)
        if dialog.exec_() == QPrintDialog.Accepted:
            try:
                painter = QPainter(printer)
                rect = printer.pageRect()
                # Scale card to fit the printer page with a small margin
                scaled_pixmap = self.card_pixmap.scaled(
                    int(rect.width() * 0.95), int(rect.height() * 0.95),
                    Qt.KeepAspectRatio, Qt.SmoothTransformation
                )
                x = (rect.width() - scaled_pixmap.width()) // 2
                y = (rect.height() - scaled_pixmap.height()) // 2
                painter.drawPixmap(x, y, scaled_pixmap)
                painter.end()
            except Exception as e:
                QMessageBox.critical(self, "Print Error",
                                     f"Failed to print the card:\n{e}")


class BatchQRCodeDialog(QDialog):
    """Dialog for batch QR card generation with an improved folder creation workflow."""

    def __init__(self, user_manager, settings, parent=None):
        super().__init__(parent)
        self.user_manager = user_manager
        self.settings = settings
        self.all_users = self.user_manager.get_all_users()
        self.setWindowTitle("Batch QR Code Generator")
        self.setMinimumSize(500, 320)  # Slightly taller for better layout
        self.init_ui()

    def init_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(15)
        main_layout.setContentsMargins(20, 20, 20, 20)
        header_layout = QHBoxLayout()
        header_layout.setSpacing(15)
        icon_label = QLabel()
        try:
            pixmap = QPixmap("asset/logo/logo.png")
            if not pixmap.isNull():
                icon_label.setPixmap(pixmap.scaled(
                    64, 64, Qt.KeepAspectRatio, Qt.SmoothTransformation))
        except Exception:
            pass
        header_layout.addWidget(icon_label, alignment=Qt.AlignTop)
        title_block = QVBoxLayout()
        title = QLabel("Batch QR Code Generation")
        title.setFont(QFont(FONT_FAMILY, 15, QFont.Bold))
        subtitle = QLabel(
            "Generate printable QR cards and save them to a new, named folder.")
        subtitle.setWordWrap(True)
        title_block.addWidget(title)
        title_block.addWidget(subtitle)
        header_layout.addLayout(title_block)
        main_layout.addLayout(header_layout)

        # Step 1: Class Selection
        selection_group = QGroupBox("Step 1: Select the Class")
        selection_group.setFont(QFont(FONT_FAMILY, 10, QFont.Bold))
        selection_layout = QVBoxLayout(selection_group)
        self.class_combo = QComboBox()
        self.class_combo.setFont(QFont(FONT_FAMILY, 10))
        self.class_combo.addItems(
            ["All Classes"] + sorted(self.user_manager.get_classes()))
        self.class_combo.currentIndexChanged.connect(self.update_user_count)
        selection_layout.addWidget(self.class_combo)
        self.user_count_label = QLabel()
        self.user_count_label.setFont(QFont(FONT_FAMILY, 9, QFont.Bold))
        self.user_count_label.setAlignment(Qt.AlignCenter)
        self.user_count_label.setStyleSheet(
            "padding: 5px; background-color: #EAF0FA; border-radius: 4px;")
        selection_layout.addWidget(self.user_count_label)
        main_layout.addWidget(selection_group)
        main_layout.addStretch()

        # Action Buttons
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        cancel_button = QPushButton("Cancel")
        cancel_button.clicked.connect(self.reject)
        generate_button = QPushButton("Step 2: Create Folder & Generate...")
        generate_button.setFont(QFont(FONT_FAMILY, 10, QFont.Bold))
        generate_button.setStyleSheet(
            f"background-color: {BUTTON_COLOR}; color: white; padding: 8px 15px;")
        generate_button.clicked.connect(self.start_generation_process)
        button_layout.addWidget(cancel_button)
        button_layout.addWidget(generate_button)
        main_layout.addLayout(button_layout)
        self.update_user_count()

    def update_user_count(self):
        selected_class = self.class_combo.currentText()
        count = len(self.all_users) if selected_class == "All Classes" else sum(
            1 for user in self.all_users if user.get('class') == selected_class
        )
        self.user_count_label.setText(f"Found {count} User(s) in Selection")

    def start_generation_process(self):
        selected_class_text = self.class_combo.currentText()
        users_to_process = [
            u for u in self.all_users if selected_class_text == "All Classes" or u.get('class') == selected_class_text
        ]
        if not users_to_process:
            QMessageBox.warning(
                self, "No Users Found", f"No users were found for the class '{selected_class_text}'.")
            return

        # Step 1: Ask for a base directory to create the new folder in.
        base_dir = self.settings.get("qr_save_dir", "library_data/qr_codes")
        parent_folder = QFileDialog.getExistingDirectory(
            self, "Select a Parent Folder (A new folder will be created inside)", base_dir)
        if not parent_folder:
            return  # User canceled

        # Step 2: Ask for the new folder's name, providing a smart default.
        safe_class_name = selected_class_text.replace(" ", "_")
        today_date = datetime.date.today().strftime('%Y-%m-%d')
        default_folder_name = f"QR_Cards_{safe_class_name}_{today_date}"

        folder_name, ok = QInputDialog.getText(
            self, "Name Your New Folder", "Enter a name for the new folder:", QLineEdit.Normal, default_folder_name)
        if not ok or not folder_name.strip():
            return  # User canceled or entered an empty name

        # Step 3: Create the full path and the directory.
        folder_path = os.path.join(parent_folder, folder_name.strip())
        try:
            os.makedirs(folder_path, exist_ok=True)
        except OSError as e:
            QMessageBox.critical(self, "Folder Creation Failed",
                                 f"Could not create the specified folder:\n{e}")
            return

        # --- Start the Progress Dialog and Generation Loop ---
        progress = QProgressDialog(
            "Generating QR cards...", "Cancel", 0, len(users_to_process), self)
        progress.setWindowTitle("Processing...")
        progress.setWindowModality(Qt.WindowModal)
        progress.setMinimumDuration(0)  # Show dialog immediately

        success_count, failed_users = 0, []
        for i, user in enumerate(users_to_process):
            progress.setValue(i)
            # Improve progress text
            progress.setLabelText(
                f"({i+1}/{len(users_to_process)}) Generating for:\n{user.get('first_name', '')} {user.get('last_name', '')}")
            if progress.wasCanceled():
                break

            try:
                card_pixmap = create_qr_card_pixmap(user)
                if card_pixmap.isNull():
                    raise RuntimeError("The generated card image is invalid.")

                # Sanitize filename
                base_name = f"{user.get('id', 'unknown')}_{user.get('first_name', '')}_{user.get('last_name', '')}".replace(
                    ' ', '_')
                safe_name = re.sub(r'[\\/:*?"<>|]', '_', base_name)
                output_path = os.path.join(folder_path, f"{safe_name}.png")

                if not card_pixmap.save(output_path, "PNG"):
                    raise IOError(
                        f"Failed to save file. Check disk space or permissions.")

                success_count += 1
            except Exception as e:
                logging.error(
                    f"Failed to generate card for user {user.get('id', 'N/A')}: {e}")
                failed_users.append(
                    f"{user.get('first_name', 'Unknown')} {user.get('last_name', '')} (ID: {user.get('id', 'N/A')}) - Reason: {str(e)}")

        progress.setValue(len(users_to_process))

        # --- Final Report with "Show Folder" button ---
        msg_box = QMessageBox(self)
        msg_box.setWindowTitle("Batch Generation Complete")

        if success_count > 0:
            msg_box.setIcon(QMessageBox.Information)
            message = f"<b>Successfully generated {success_count} of {len(users_to_process)} QR card(s).</b>"
            if failed_users:
                message += f"\n\nHowever, {len(failed_users)} card(s) failed to generate. Please check the log for details."

            msg_box.setText(message)
            msg_box.setInformativeText(f"Files were saved in:\n{folder_path}")

            show_folder_button = msg_box.addButton(
                "📂 Show Folder", QMessageBox.ActionRole)
            msg_box.addButton(QMessageBox.Close)

            msg_box.exec_()

            if msg_box.clickedButton() == show_folder_button:
                QDesktopServices.openUrl(QUrl.fromLocalFile(folder_path))
        else:
            msg_box.setIcon(QMessageBox.Critical)
            error_details = "\n".join(failed_users[:5])  # Show first 5 errors
            msg_box.setText(f"<b>Batch Generation Failed</b>")
            msg_box.setInformativeText(
                f"Could not save any QR cards. Common causes include disk space issues or missing user IDs.\n\n<b>Some errors:</b>\n{error_details}")
            msg_box.exec_()

        # Save the parent directory for next time
        self.settings["qr_save_dir"] = parent_folder

# Log Viewer Dialog


class LogViewerDialog(QDialog):
    def __init__(self, log_manager, user_manager, settings, parent=None):
        super().__init__(parent)
        self.log_manager = log_manager
        self.user_manager = user_manager
        self.settings = settings
        self.setWindowTitle("Log Viewer - Detailed Attendance Records")
        self.setMinimumSize(1000, 700)
        self.logs_cache = []
        self.init_ui()
        self.load_logs()

    def init_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(12)
        main_layout.setContentsMargins(15, 15, 15, 15)

        # Date Range & Actions
        top_layout = QHBoxLayout()
        top_layout.addWidget(QLabel("From:"))
        self.start_date_edit = QDateEdit(QDate.currentDate().addDays(-7))
        self.start_date_edit.setCalendarPopup(True)
        self.start_date_edit.dateChanged.connect(self.load_logs)
        top_layout.addWidget(self.start_date_edit)
        top_layout.addWidget(QLabel("To:"))
        self.end_date_edit = QDateEdit(QDate.currentDate())
        self.end_date_edit.setCalendarPopup(True)
        self.end_date_edit.dateChanged.connect(self.load_logs)
        top_layout.addWidget(self.end_date_edit)

        refresh_btn = QPushButton("Refresh Logs")
        refresh_btn.clicked.connect(self.load_logs)
        export_btn = QPushButton("Export to Excel")
        export_btn.clicked.connect(self.export_logs)
        top_layout.addWidget(refresh_btn)
        top_layout.addWidget(export_btn)
        main_layout.addLayout(top_layout)

        # Search Bar
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Search by ID, Name, or Class...")
        self.search_input.textChanged.connect(self.filter_table)
        main_layout.addWidget(self.search_input)

        # Log Table
        self.log_table = QTableWidget()
        self.log_table.setColumnCount(9)
        self.log_table.setHorizontalHeaderLabels(
            ['Date', 'User ID', 'First Name', 'Last Name', 'Class', 'Gender', 'Check-ins', 'First In', 'Last Out'])
        header = self.log_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeToContents)
        header.setStretchLastSection(True)
        self.log_table.setSortingEnabled(True)
        self.log_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.log_table.setEditTriggers(QTableWidget.NoEditTriggers)
        main_layout.addWidget(self.log_table)

        # Admin Actions
        clear_btn = QPushButton("Clear All Logs (Admin Only)")
        clear_btn.setStyleSheet("background-color: #E74C3C;")
        clear_btn.clicked.connect(self.clear_all_logs)
        main_layout.addWidget(clear_btn, alignment=Qt.AlignCenter)

    def load_logs(self):
        start_date = self.start_date_edit.date().toPyDate()
        end_date = self.end_date_edit.date().toPyDate()
        if start_date > end_date:
            QMessageBox.warning(self, "Date Error",
                                "Start date cannot be after end date.")
            return

        self.log_table.setSortingEnabled(False)
        self.log_table.setRowCount(0)

        log_data = self.log_manager.get_detailed_log_data_for_date_range(
            start_date, end_date)

        processed = {}
        for user_entry in log_data:
            df = user_entry['logs']
            df['timestamp'] = pd.to_datetime(df['timestamp'])
            for date_key, group in df.groupby(df['timestamp'].dt.date):
                key = (user_entry['user_id'], date_key)
                if key not in processed:
                    processed[key] = {
                        **user_entry, 'date': date_key, 'check_ins': [], 'check_outs': []}

                processed[key]['check_ins'].extend(
                    group[group['event'] == 'Check-in']['timestamp'].tolist())
                processed[key]['check_outs'].extend(
                    group[group['event'] == 'Check-out']['timestamp'].tolist())

        self.logs_cache = list(processed.values())
        self.populate_table()
        self.log_table.setSortingEnabled(True)

    def populate_table(self):
        self.log_table.setRowCount(len(self.logs_cache))
        for row, log in enumerate(self.logs_cache):
            items = [
                log['date'].strftime('%Y-%m-%d'),
                str(log['user_id']),
                log['first_name'],
                log['last_name'],
                log.get('class', ''),
                log.get('gender', ''),
                str(len(log['check_ins'])),
                min(log['check_ins']).strftime(
                    '%H:%M:%S') if log['check_ins'] else '—',
                max(log['check_outs']).strftime(
                    '%H:%M:%S') if log['check_outs'] else '—'
            ]
            for col, text in enumerate(items):
                self.log_table.setItem(row, col, QTableWidgetItem(text))

    def filter_table(self):
        query = self.search_input.text().lower().strip()
        for row in range(self.log_table.rowCount()):
            row_text = ' '.join([self.log_table.item(
                row, col).text().lower() for col in [1, 2, 3, 4]])
            self.log_table.setRowHidden(row, query not in row_text)

    def export_logs(self):
        if not self.logs_cache:
            QMessageBox.information(
                self, "No Data", "There is no data to export.")
            return

        # Let the user choose where to save the file
        default_filename = f"Attendance_Report_{self.start_date_edit.date().toString('yyyy-MM-dd')}_to_{self.end_date_edit.date().toString('yyyy-MM-dd')}.xlsx"
        filename, _ = QFileDialog.getSaveFileName(
            self, "Save Excel Report", default_filename, "Excel Files (*.xlsx)")

        if not filename:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = "Detailed Attendance Report"

            # --- Page Setup for Printing ---
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            ws.page_setup.fitToPage = True
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.page_margins.left = 0.5
            ws.page_margins.right = 0.5
            ws.page_margins.top = 0.75
            ws.page_margins.bottom = 0.75

            # --- Define Styles ---
            # Using Khmer OS Siemreap for broad compatibility
            main_header_font = Font(name='Khmer OS Muol', size=14, bold=True)
            subtitle_font = Font(name='Khmer OS Siemreap', size=11)
            table_header_font = Font(
                name='Khmer OS Siemreap', size=10, bold=True)
            body_font = Font(name='Khmer OS Siemreap', size=10)
            footer_font = Font(name='Khmer OS Siemreap', size=10, italic=True)

            center_align = Alignment(
                horizontal='center', vertical='center', wrap_text=True)
            left_align = Alignment(horizontal='left', vertical='center')
            thin_border = Border(left=Side(style='thin'), right=Side(
                style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            # --- Column Headers (in Khmer) ---
            columns = [
                'ល.រ.', 'កាលបរិច្ឆេទ', 'អត្តលេខ', 'នាមត្រកូល', 'នាមខ្លួន', 'ថ្នាក់',
                'ភេទ', 'ចំនួនដងចូល', 'ពេលចូលដំបូង', 'ពេលចេញចុងក្រោយ'
            ]
            num_cols = len(columns)
            end_col_letter = get_column_letter(num_cols)

            # --- Main Title and Subtitle ---
            ws.merge_cells(f'A1:{end_col_letter}1')
            ws['A1'].value = "របាយការណ៍វត្តមានសិស្សលម្អិត"
            ws['A1'].font = main_header_font
            ws['A1'].alignment = center_align

            start_date_str = self.start_date_edit.date().toString("dd/MM/yyyy")
            end_date_str = self.end_date_edit.date().toString("dd/MM/yyyy")
            ws.merge_cells(f'A2:{end_col_letter}2')
            ws['A2'].value = f"ពីកាលបរិច្ឆេទ {start_date_str} ដល់ {end_date_str}"
            ws['A2'].font = subtitle_font
            ws['A2'].alignment = center_align
            ws.row_dimensions[1].height = 30
            ws.row_dimensions[2].height = 25

            # --- Table Headers ---
            header_row_index = 4
            for col_num, header_title in enumerate(columns, 1):
                cell = ws.cell(row=header_row_index, column=col_num)
                cell.value = header_title
                cell.font = table_header_font
                cell.alignment = center_align
                cell.border = thin_border
            ws.row_dimensions[header_row_index].height = 22

            # --- Populate Data Rows ---
            current_row = header_row_index + 1
            # Sort data by date, then by user ID for a logical flow
            sorted_logs = sorted(self.logs_cache, key=lambda log: (
                log['date'], log['user_id']))

            for index, log in enumerate(sorted_logs, 1):
                first_in_str = min(log['check_ins']).strftime(
                    '%H:%M:%S') if log['check_ins'] else '—'
                last_out_str = max(log['check_outs']).strftime(
                    '%H:%M:%S') if log['check_outs'] else '—'

                row_data = [
                    index,
                    log['date'].strftime('%d-%b-%Y'),
                    log['user_id'],
                    log['first_name'],
                    log['last_name'],
                    log.get('class', ''),
                    log.get('gender', ''),
                    len(log['check_ins']),
                    first_in_str,
                    last_out_str
                ]
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=col_num)
                    cell.value = value
                    cell.font = body_font
                    cell.border = thin_border
                    # Center align all data for a clean look
                    cell.alignment = center_align
                current_row += 1

            # --- Auto-adjust Column Widths ---
            column_widths = {}
            for row in ws.iter_rows(min_row=header_row_index, max_row=current_row - 1):
                for cell in row:
                    col_letter = get_column_letter(cell.column)
                    # Use a multiplier for Khmer fonts which are wider
                    length = len(str(cell.value)) * 1.4
                    if col_letter not in column_widths or length > column_widths[col_letter]:
                        column_widths[col_letter] = length

            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = max(
                    width + 2, 12)  # Add padding, with a minimum width
            # Fixed width for the index column
            ws.column_dimensions['A'].width = 8

            # --- Footer Section ---
            footer_row_index = current_row + 1
            today = datetime.datetime.now()
            # Format date using Khmer Unicode for day and month
            khmer_day = "ថ្ងៃទី" + \
                str(today.day).translate(
                    str.maketrans("0123456789", "០១២៣៤៥៦៧៨៩"))
            khmer_months = ["មករា", "កុម្ភៈ", "មីនា", "មេសា", "ឧសភា",
                            "មិថុនា", "កក្កដា", "សីហា", "កញ្ញា", "តុលា", "វិច្ឆិកា", "ធ្នូ"]
            khmer_month_str = "ខែ" + khmer_months[today.month - 1]
            khmer_year = "ឆ្នាំ" + \
                str(today.year).translate(
                    str.maketrans("0123456789", "០១២៣៤៥៦៧៨៩"))

            # Place "Phnom Penh, Date..." in the footer
            footer_date_text = f"វិទ្យាល័យសម្តេចឪ សម្តេចម៉ែ, {khmer_day} {khmer_month_str} {khmer_year}"
            ws.merge_cells(
                f'{get_column_letter(num_cols-3)}{footer_row_index}:{end_col_letter}{footer_row_index}')
            footer_cell = ws.cell(row=footer_row_index, column=num_cols-3)
            footer_cell.value = footer_date_text
            footer_cell.font = footer_font
            footer_cell.alignment = center_align

            # Add a signature line
            signature_row_index = footer_row_index + 1
            ws.merge_cells(
                f'{get_column_letter(num_cols-3)}{signature_row_index}:{end_col_letter}{signature_row_index}')
            signature_cell = ws.cell(
                row=signature_row_index, column=num_cols-3)
            signature_cell.value = "ហត្ថលេខាបណ្ណារក្ស"
            signature_cell.font = table_header_font
            signature_cell.alignment = center_align

            # --- Save the Workbook ---
            wb.save(filename)
            QMessageBox.information(
                self, "Export Successful", f"The report has been saved to:\n{filename}")

        except Exception as e:
            QMessageBox.critical(self, "Export Failed",
                                 f"An unexpected error occurred:\n{e}")

    def clear_all_logs(self):
        if QMessageBox.question(self, "Confirm Clear", "⚠️ Delete ALL attendance logs permanently?", QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
            try:
                self.log_manager.clear_logs()
                self.load_logs()
                QMessageBox.information(
                    self, "Success", "All logs have been cleared.")
            except Exception as e:
                QMessageBox.critical(
                    self, "Error", f"Failed to clear logs:\n{e}")

# Replace the entire existing ExportClassDialog class with this one.


class ExportClassDialog(QDialog):
    def __init__(self, user_manager, log_manager, settings, parent=None):
        super().__init__(parent)
        self.user_manager = user_manager
        self.log_manager = log_manager
        self.settings = settings
        self.setWindowTitle("Export Class Data")
        self.setMinimumSize(500, 500)
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(GROUP_SPACING)
        layout.setContentsMargins(
            DIALOG_PADDING, DIALOG_PADDING, DIALOG_PADDING, DIALOG_PADDING)
        header = QLabel("Export Class Data")
        header.setFont(QFont(UI_FONT_FAMILY, LARGE_FONT_SIZE, QFont.Bold))
        header.setAlignment(Qt.AlignCenter)
        layout.addWidget(header)
        class_layout = QHBoxLayout()
        self.class_combo = QComboBox()
        classes = self.user_manager.get_classes()
        if not classes:
            QMessageBox.warning(self, "No Classes",
                                "No classes available to export.")
            self.reject()
            return
        self.class_combo.addItems(classes)
        class_layout.addWidget(QLabel("Select Class:"))
        class_layout.addWidget(self.class_combo)
        layout.addLayout(class_layout)
        mode_layout = QHBoxLayout()
        self.by_day_radio = QRadioButton("By Day")
        self.by_day_radio.setChecked(True)
        self.by_day_radio.toggled.connect(self.toggle_mode)
        mode_layout.addWidget(self.by_day_radio)
        self.by_month_radio = QRadioButton("By Month")
        self.by_month_radio.toggled.connect(self.toggle_mode)
        mode_layout.addWidget(self.by_month_radio)
        self.by_multi_month_radio = QRadioButton("By Multi-Month")
        self.by_multi_month_radio.toggled.connect(self.toggle_mode)
        mode_layout.addWidget(self.by_multi_month_radio)
        layout.addLayout(mode_layout)
        self.day_widget = QWidget()
        self.day_layout = QHBoxLayout()
        self.day_edit = QDateEdit(QDate.currentDate())
        self.day_edit.setCalendarPopup(True)
        self.day_layout.addWidget(QLabel("Select Date:"))
        self.day_layout.addWidget(self.day_edit)
        self.day_widget.setLayout(self.day_layout)
        layout.addWidget(self.day_widget)
        self.month_widget = QWidget()
        self.month_layout = QHBoxLayout()
        self.month_combo = QComboBox()
        months = ["January", "February", "March", "April", "May", "June",
                  "July", "August", "September", "October", "November", "December"]
        self.month_combo.addItems(months)
        self.month_combo.setCurrentIndex(QDate.currentDate().month() - 1)
        self.year_combo = QComboBox()
        current_year = QDate.currentDate().year()
        years = [str(y) for y in range(current_year - 5, current_year + 2)]
        self.year_combo.addItems(years)
        self.year_combo.setCurrentText(str(current_year))
        self.month_layout.addWidget(QLabel("Select Month:"))
        self.month_layout.addWidget(self.month_combo)
        self.month_layout.addWidget(self.year_combo)
        self.month_widget.setLayout(self.month_layout)
        self.month_widget.setVisible(False)
        layout.addWidget(self.month_widget)
        self.multi_month_widget = QWidget()
        self.multi_month_layout = QVBoxLayout()
        self.multi_month_list = QListWidget()
        self.multi_month_list.setSelectionMode(QListWidget.MultiSelection)
        for month in months:
            item = QListWidgetItem(month)
            self.multi_month_list.addItem(item)
        self.multi_month_year_combo = QComboBox()
        self.multi_month_year_combo.addItems(years)
        self.multi_month_year_combo.setCurrentText(str(current_year))
        self.multi_month_layout.addWidget(
            QLabel("Select Months (Hold Ctrl for multiple):"))
        self.multi_month_layout.addWidget(self.multi_month_list)
        self.multi_month_layout.addWidget(QLabel("Select Year:"))
        self.multi_month_layout.addWidget(self.multi_month_year_combo)
        self.multi_month_widget.setLayout(self.multi_month_layout)
        self.multi_month_widget.setVisible(False)
        layout.addWidget(self.multi_month_widget)

        signature_layout = QHBoxLayout()
        signature_layout.addWidget(QLabel("<b>Signature Start Column:</b>"))
        self.signature_col_spinbox = QSpinBox()
        self.signature_col_spinbox.setRange(1, 50)
        # Default to column E for daily report
        self.signature_col_spinbox.setValue(5)
        signature_layout.addWidget(self.signature_col_spinbox)
        layout.addLayout(signature_layout)

        button_layout = QHBoxLayout()
        export_button = QPushButton("Export Data")
        export_button.clicked.connect(self.export_class_data)
        export_button.setStyleSheet(
            f"background-color: {BUTTON_COLOR}; color: {BUTTON_TEXT_COLOR}; padding: 8px; border-radius: 6px;")
        button_layout.addWidget(export_button)
        close_button = QPushButton("Close")
        close_button.clicked.connect(self.accept)
        close_button.setStyleSheet(
            f"background-color: #E74C3C; color: white; padding: 8px; border-radius: 6px;")
        button_layout.addWidget(close_button)
        layout.addLayout(button_layout)
        self.setLayout(layout)
        self.apply_theme()

    def apply_theme(self):
        self.setStyleSheet(f"""
            QDialog {{ background-color: {THEME_BG_COLOR}; color: {TEXT_COLOR}; border-radius: 10px; }}
            QPushButton:hover {{ background-color: {BUTTON_HOVER_COLOR}; }}
            QComboBox, QDateEdit, QListWidget, QSpinBox {{ padding: 5px; border: 1px solid #BDC3C7; border-radius: 5px; background-color: {THEME_BG_COLOR}; color: {TEXT_COLOR}; }}
            QRadioButton {{ color: {TEXT_COLOR}; }}
        """)

    def toggle_mode(self):
        is_day_mode = self.by_day_radio.isChecked()
        self.day_widget.setVisible(is_day_mode)
        self.month_widget.setVisible(self.by_month_radio.isChecked())
        self.multi_month_widget.setVisible(
            self.by_multi_month_radio.isChecked())
        # Adjust default signature column based on report type
        if is_day_mode:
            # Column E is reasonable for daily report
            self.signature_col_spinbox.setValue(5)
        else:
            # Column V for wider monthly report
            self.signature_col_spinbox.setValue(22)

    def export_class_data(self):
        selected_class = self.class_combo.currentText()
        if not selected_class:
            QMessageBox.warning(self, "Selection Error",
                                "Please select a class to export")
            return
        if self.by_day_radio.isChecked():
            selected_date = self.day_edit.date().toPyDate()
            self.export_by_day(selected_class, selected_date)
        elif self.by_month_radio.isChecked():
            month = self.month_combo.currentIndex() + 1
            year = int(self.year_combo.currentText())
            self.export_by_month(selected_class, year, month)
        else:
            selected_months = [item.text()
                               for item in self.multi_month_list.selectedItems()]
            year = int(self.multi_month_year_combo.currentText())
            if not selected_months:
                QMessageBox.warning(self, "Selection Error",
                                    "Please select at least one month")
                return
            self.export_by_multi_month(selected_class, year, selected_months)

    def export_by_day(self, selected_class, selected_date):
        from openpyxl.styles import Border, Side, Alignment, Font
        from openpyxl.utils import get_column_letter
        from openpyxl import Workbook
        import os
        from PyQt5.QtWidgets import QMessageBox
        from datetime import datetime, date
        import pandas as pd

        log_data = self.log_manager.get_detailed_log_data_for_date_range(
            selected_date, selected_date)
        class_users_data = [u for u in log_data if u.get(
            'class') == selected_class]

        if not class_users_data:
            QMessageBox.information(
                self, "No Data",
                f"No attendance data for class {selected_class} on {selected_date.strftime('%Y-%m-%d')}"
            )
            return

        exported_data = []
        female_present_count = 0
        total_present_count = 0

        def parse_time(ts):
            if pd.isna(ts) or ts is None:
                return None
            if isinstance(ts, datetime):
                return ts
            try:
                return pd.to_datetime(ts)
            except (ValueError, TypeError):
                return None

        for user_data in class_users_data:
            logs_df = user_data.get('logs')
            if logs_df is None or logs_df.empty:
                continue

            check_ins = logs_df[logs_df['event'] == 'Check-in'].copy()
            check_outs = logs_df[logs_df['event'] == 'Check-out'].copy()

            check_in_count = len(check_ins)
            if check_in_count == 0:
                continue

            total_present_count += 1
            if user_data.get('gender', '').lower() in ['female', 'ស្រី']:
                female_present_count += 1

            check_ins['parsed_time'] = check_ins['timestamp'].apply(parse_time)
            check_outs['parsed_time'] = check_outs['timestamp'].apply(
                parse_time)
            first_check_in_time = check_ins['parsed_time'].dropna().min()
            last_check_out_time = check_outs['parsed_time'].dropna().max()
            check_in_str = first_check_in_time.strftime(
                '%H:%M:%S') if pd.notna(first_check_in_time) else 'N/A'
            check_out_str = last_check_out_time.strftime(
                '%H:%M:%S') if pd.notna(last_check_out_time) else 'N/A'
            name = f"{user_data.get('first_name', '')} {user_data.get('last_name', '')}".strip(
            ) or "Unknown User"

            exported_data.append({
                'ID': user_data['user_id'], 'Name': name, 'Sex': str(user_data.get('gender', 'N/A')),
                'Class': str(user_data.get('class', 'N/A')), 'Study Year': str(user_data.get('study_year', 'N/A')),
                'Status': str(check_in_count), 'Check-in Time': check_in_str, 'Check-out Time': check_out_str
            })

        wb = Workbook()
        ws = wb.active
        ws.title = "Daily Report"
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

        thin_border = Border(left=Side(style='thin'), right=Side(
            style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        columns = ['ល.រ', 'អត្ថលេខ', 'ឈ្មោះ', 'ភេទ', 'ថ្នាក់',
                   'ឆ្នាំសិក្សា', 'ចំនួនដងចូល', 'ពេលចូល', 'ពេលចេញ']
        num_cols = len(columns)
        end_col_letter = get_column_letter(num_cols)

        ws.merge_cells(f'A1:{end_col_letter}1')
        ws['A1'].value = "ព្រះរាជាណាចក្រកម្ពុជា"
        ws['A1'].font = Font(name='Khmer OS Muol', size=12, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells(f'A2:{end_col_letter}2')
        ws['A2'].value = "ជាតិ សាសនា ព្រះមហាក្សត្រ"
        ws['A2'].font = Font(name='Khmer OS Muol', size=11, bold=True)
        ws['A2'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A4:F4')
        ws['A4'].value = "វិទ្យាល័យសម្តេចឪ សម្តេចម៉ែ"
        ws['A4'].font = Font(name='Khmer OS Muol', size=10, bold=True)

        ws.merge_cells(f'A6:{end_col_letter}6')
        ws['A6'].value = "របាយការណ៍វត្តមានប្រចាំថ្ងៃ"
        ws['A6'].font = Font(name='Khmer OS Muol', size=11, bold=True)
        ws['A6'].alignment = Alignment(horizontal='center', vertical='center')

        khmer_months = ["មករា", "កុម្ភៈ", "មីនា", "មេសា", "ឧសភា",
                        "មិថុនា", "កក្កដា", "សីហា", "កញ្ញា", "តុលា", "វិច្ឆិកា", "ធ្នូ"]
        date_str = f"សម្រាប់ថ្ងៃទី {selected_date.day} ខែ {khmer_months[selected_date.month - 1]} ឆ្នាំ {selected_date.year}"
        ws.merge_cells(f'A7:{end_col_letter}7')
        ws['A7'].value = date_str
        ws['A7'].font = Font(name='Khmer OS Siemreap', size=10)
        ws['A7'].alignment = Alignment(horizontal='center', vertical='center')

        header_row = 9
        for col_num, header_title in enumerate(columns, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.value = header_title
            cell.font = Font(name='Khmer OS Siemreap', size=9, bold=True)
            cell.alignment = Alignment(
                horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        ws.row_dimensions[header_row].height = 20

        for i, user_info in enumerate(sorted(exported_data, key=lambda x: x['ID']), 1):
            row_data = [i, user_info['ID'], user_info['Name'], user_info['Sex'], user_info['Class'],
                        user_info['Study Year'], user_info['Status'], user_info['Check-in Time'], user_info['Check-out Time']]
            for j, value in enumerate(row_data):
                cell = ws.cell(row=header_row + i, column=j + 1, value=value)
                cell.font = Font(name='Khmer OS Siemreap', size=9)
                cell.alignment = Alignment(
                    horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
            ws.row_dimensions[header_row + i].height = 16

        summary_row = header_row + len(exported_data) + 2
        ws['A' + str(summary_row)
           ].value = f"សរុបសិស្សមានវត្តមាន: {total_present_count}"
        ws['A' + str(summary_row)
           ].font = Font(name='Khmer OS Siemreap', size=10, bold=True)
        ws['A' + str(summary_row + 1)
           ].value = f"សិស្សស្រីមានវត្តមាន: {female_present_count}"
        ws['A' + str(summary_row + 1)
           ].font = Font(name='Khmer OS Siemreap', size=10, bold=True)

        # MODIFICATION START: Added footer to daily report
        footer_start_row = summary_row + 2
        lunar_start_col_index = self.signature_col_spinbox.value()
        lunar_end_col_index = lunar_start_col_index + \
            4  # A shorter merge for a narrower report

        lunar_date_cell = ws.cell(
            row=footer_start_row, column=lunar_start_col_index)
        lunar_date_cell.value = "ធ្វើនៅ ថ្ងៃ.......... ខែ.......... ឆ្នាំ...............ស័ក ព.ស.........."
        lunar_date_cell.font = Font(name='Khmer OS Siemreap')
        ws.merge_cells(start_row=footer_start_row, start_column=lunar_start_col_index,
                       end_row=footer_start_row, end_column=lunar_end_col_index)
        lunar_date_cell.alignment = Alignment(horizontal='center')

        today = date.today()
        gregorian_date_str = f"ត្រូវនឹងថ្ងៃទី {today.day} ខែ {khmer_months[today.month - 1]} ឆ្នាំ {today.year}"
        gregorian_date_cell = ws.cell(
            row=footer_start_row + 1, column=lunar_start_col_index)
        gregorian_date_cell.value = gregorian_date_str
        gregorian_date_cell.font = Font(name='Khmer OS Siemreap')
        ws.merge_cells(start_row=footer_start_row + 1, start_column=lunar_start_col_index,
                       end_row=footer_start_row + 1, end_column=lunar_end_col_index)
        gregorian_date_cell.alignment = Alignment(horizontal='center')

        signature_cell = ws.cell(
            row=footer_start_row + 3, column=lunar_start_col_index)
        signature_cell.value = "ហត្ថលេខាបណ្ណារ័ក្ស"
        signature_cell.font = Font(
            name='Khmer OS Siemreap', size=10, bold=False)
        ws.merge_cells(start_row=footer_start_row + 3, start_column=lunar_start_col_index,
                       end_row=footer_start_row + 3, end_column=lunar_end_col_index)
        signature_cell.alignment = Alignment(horizontal='center')
        # MODIFICATION END

        col_widths = [6, 10, 22, 8, 10, 12, 12, 12, 12]
        for i, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        filename = f"daily_class_{selected_class}_attendance_{selected_date.strftime('%Y%m%d')}.xlsx"
        filepath = os.path.join(self.settings["qr_save_dir"], filename)
        try:
            wb.save(filepath)
            QMessageBox.information(
                self, "Success", f"Daily report for class {selected_class} exported to: {filepath}")
        except Exception as e:
            QMessageBox.critical(
                self, "Error", f"Failed to save daily report: {e}")

    def export_by_month(self, selected_class, year, month):
        from openpyxl.styles import Border, Side
        from openpyxl.utils import get_column_letter, column_index_from_string
        from datetime import date, timedelta
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        import os
        from PyQt5.QtWidgets import QMessageBox

        start_date = date(year, month, 1)
        last_day = (date(year, month + 1, 1) -
                    timedelta(days=1)).day if month < 12 else 31
        end_date = date(year, month, last_day)

        log_data = self.log_manager.get_detailed_log_data_for_date_range(
            start_date, end_date)
        class_users = [user for user in self.user_manager.get_all_users(
        ) if user.get('class') == selected_class]
        if not class_users:
            QMessageBox.information(
                self, "No Users", f"No users found in class {selected_class}.")
            return

        user_attendance = {}
        total_present_person_days = 0
        female_present_person_days = 0

        for user in class_users:
            user_id = user['id']
            user_attendance[user_id] = {
                day: 0 for day in range(1, last_day + 1)}
            user_attendance[user_id]['user_info'] = user

        for user_log_data in log_data:
            user_id = user_log_data.get('user_id')
            if user_id not in user_attendance:
                continue
            for _, log in user_log_data['logs'].iterrows():
                event_time = pd.to_datetime(log['timestamp'])
                event_date = event_time.date()
                if event_date.year == year and event_date.month == month and log['event'] == 'Check-in':
                    day = event_date.day
                    user_attendance[user_id][day] += 1

        exported_data = []
        for user_id, attendance_data in user_attendance.items():
            user_info_dict = attendance_data['user_info']
            name = f"{user_info_dict.get('first_name', '')} {user_info_dict.get('last_name', '')}".strip(
            ) or "Unknown User"
            daily_counts = [attendance_data[day]
                            for day in range(1, last_day + 1)]

            total_checkins_for_user = sum(daily_counts)
            days_attended_by_user = sum(
                1 for count in daily_counts if count > 0)

            total_present_person_days += days_attended_by_user
            if user_info_dict.get('gender', '').lower() in ['female', 'ស្រី']:
                female_present_person_days += days_attended_by_user

            user_info_export = {
                'ID': user_id, 'Name': name, 'Sex': str(user_info_dict.get('gender', 'N/A')),
                'Class': str(user_info_dict.get('class', 'N/A')), 'Study Year': str(user_info_dict.get('study_year', DEFAULT_STUDY_YEAR)),
                'Total Attendance': total_checkins_for_user
            }
            for day in range(1, last_day + 1):
                count = attendance_data[day]
                user_info_export[f'{day}'] = count if count > 0 else ''
            exported_data.append(user_info_export)

        wb = Workbook()
        ws = wb.active
        ws.title = "Monthly Attendance Report"
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

        thin_border = Border(left=Side(style='thin'), right=Side(
            style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        columns = ['ល.រ', 'អត្ថលេខ', 'ឈ្មោះ', 'ភេទ', 'ថ្នាក់', 'ឆ្នាំសិក្សា'] + \
            [f'{day}' for day in range(1, last_day + 1)] + ['សរុប']
        num_cols = len(columns)
        end_col_letter = get_column_letter(num_cols)

        ws.merge_cells(f'A1:{end_col_letter}1')
        ws['A1'].value = "ព្រះរាជាណាចក្រកម្ពុជា"
        ws['A1'].font = Font(name='Khmer OS Muol', size=12, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells(f'A2:{end_col_letter}2')
        ws['A2'].value = "ជាតិ សាសនា ព្រះមហាក្សត្រ"
        ws['A2'].font = Font(name='Khmer OS Muol', size=11, bold=True)
        ws['A2'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A4:F4')
        ws['A4'].value = "វិទ្យាល័យសម្តេចឪ សម្តេចម៉ែ"
        ws['A4'].font = Font(name='Khmer OS Muol', size=10, bold=True)

        khmer_months = ["មករា", "កុម្ភៈ", "មីនា", "មេសា", "ឧសភា",
                        "មិថុនា", "កក្កដា", "សីហា", "កញ្ញា", "តុលា", "វិច្ឆិកា", "ធ្នូ"]
        khmer_month_name = khmer_months[month - 1]
        ws.merge_cells(f'A5:{end_col_letter}5')
        ws['A5'].value = f"របាយការណ៍វត្តមានប្រចាំខែ {khmer_month_name}"
        ws['A5'].font = Font(name='Khmer OS Muol', size=11, bold=True)
        ws['A5'].alignment = Alignment(horizontal='center')
        ws.merge_cells(f'A6:{end_col_letter}6')
        ws['A6'].value = f"ថ្នាក់ទី {selected_class}"
        ws['A6'].font = Font(name='Khmer OS Muol', size=11, bold=True)
        ws['A6'].alignment = Alignment(horizontal='center')

        header_row = 8
        for col_num, header_title in enumerate(columns, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.value = header_title
            cell.font = Font(name='Khmer OS Siemreap', bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        exported_data = sorted(exported_data, key=lambda x: x['ID'])
        for i, user_info in enumerate(exported_data, 1):
            row_data = [i, user_info['ID'], user_info['Name'], user_info['Sex'], user_info['Class'], user_info['Study Year']] + \
                       [user_info.get(f'{day}', '') for day in range(
                           1, last_day + 1)] + [user_info['Total Attendance']]
            for j, value in enumerate(row_data):
                cell = ws.cell(row=header_row + i, column=j + 1, value=value)
                cell.font = Font(name='Khmer OS Siemreap')
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

        summary_row = header_row + len(exported_data) + 2
        ws['A' + str(summary_row)
           ].value = f"សរុបវត្តមានសិស្ស (គិតជាថ្ងៃ): {total_present_person_days}"
        ws['A' + str(summary_row)
           ].font = Font(name='Khmer OS Siemreap', bold=True)
        ws['A' + str(summary_row + 1)
           ].value = f"សរុបវត្តមានសិស្សស្រី (គិតជាថ្ងៃ): {female_present_person_days}"
        ws['A' + str(summary_row + 1)
           ].font = Font(name='Khmer OS Siemreap', bold=True)

        footer_start_row = summary_row + 2
        lunar_start_col_index = self.signature_col_spinbox.value()
        lunar_end_col_index = lunar_start_col_index + 15

        lunar_date_cell = ws.cell(
            row=footer_start_row, column=lunar_start_col_index)
        lunar_date_cell.value = "ធ្វើនៅ ថ្ងៃ.......... ............. ខែ.......... ឆ្នាំ.......... .....ស័ក ព.ស.........."
        lunar_date_cell.font = Font(name='Khmer OS Siemreap')
        ws.merge_cells(start_row=footer_start_row, start_column=lunar_start_col_index,
                       end_row=footer_start_row, end_column=lunar_end_col_index)
        lunar_date_cell.alignment = Alignment(horizontal='center')

        today = date.today()
        gregorian_date_str = f"ត្រូវនឹងថ្ងៃទី {today.day} ខែ {khmer_months[today.month - 1]} ឆ្នាំ {today.year}"
        gregorian_date_cell = ws.cell(
            row=footer_start_row + 1, column=lunar_start_col_index)
        gregorian_date_cell.value = gregorian_date_str
        gregorian_date_cell.font = Font(name='Khmer OS Siemreap')
        ws.merge_cells(start_row=footer_start_row + 1, start_column=lunar_start_col_index,
                       end_row=footer_start_row + 1, end_column=lunar_end_col_index)
        gregorian_date_cell.alignment = Alignment(horizontal='center')

        signature_cell = ws.cell(
            row=footer_start_row + 3, column=lunar_start_col_index)
        signature_cell.value = "ហត្ថលេខាបណ្ណារ័ក្ស"
        signature_cell.font = Font(
            name='Khmer OS Siemreap', size=10, bold=False)
        ws.merge_cells(start_row=footer_start_row + 3, start_column=lunar_start_col_index,
                       end_row=footer_start_row + 3, end_column=lunar_end_col_index)
        signature_cell.alignment = Alignment(horizontal='center')

        for col_idx, column_cells in enumerate(ws.columns, 1):
            col_letter = get_column_letter(col_idx)
            if col_idx == 1:
                ws.column_dimensions[col_letter].width = 5
                continue
            if col_idx <= len(columns) and columns[col_idx - 1].isdigit():
                ws.column_dimensions[col_letter].width = 4
                continue
            max_length = 0
            for cell in column_cells:
                try:
                    if cell.value is not None and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

        filename = f"monthly_class_{selected_class}_attendance_{year}{month:02d}.xlsx"
        filepath = os.path.join(self.settings["qr_save_dir"], filename)
        os.makedirs(self.settings["qr_save_dir"], exist_ok=True)
        try:
            wb.save(filepath)
            QMessageBox.information(
                self, "Success", f"Class {selected_class} data for {khmer_month_name} {year} exported to: {filepath}")
        except Exception as e:
            QMessageBox.critical(
                self, "Error", f"Failed to save Excel file: {e}")

    def export_by_multi_month(self, selected_class, year, months):
        from openpyxl.utils import get_column_letter
        from datetime import date, timedelta
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side
        import os
        from PyQt5.QtWidgets import QMessageBox, QProgressDialog, QFileDialog
        import logging

        month_map = {"January": 1, "February": 2, "March": 3, "April": 4, "May": 5, "June": 6,
                     "July": 7, "August": 8, "September": 9, "October": 10, "November": 11, "December": 12}

        output_folder = QFileDialog.getExistingDirectory(
            self, "Select Folder to Save Monthly Reports", self.settings["qr_save_dir"])
        if not output_folder:
            return

        success_count, failed_months = 0, []
        progress = QProgressDialog(
            "Exporting monthly reports...", "Cancel", 0, len(months), self)
        progress.setWindowModality(Qt.WindowModal)
        progress.setMinimumDuration(0)

        for i, month_name in enumerate(months):
            progress.setValue(i)
            progress.setLabelText(f"Exporting {month_name}...")
            if progress.wasCanceled():
                break

            try:
                month = month_map[month_name]
                start_date, end_date = date(year, month, 1), date(year, month, (date(
                    year, month % 12 + 1, 1) - timedelta(days=1)).day if month < 12 else 31)
                last_day = end_date.day

                log_data = self.log_manager.get_detailed_log_data_for_date_range(
                    start_date, end_date)
                class_users = [user for user in self.user_manager.get_all_users(
                ) if user.get('class') == selected_class]
                if not class_users:
                    logging.warning(
                        f"No users in class {selected_class} for {month_name}, skipping.")
                    continue

                user_attendance, total_present_person_days, female_present_person_days = {}, 0, 0

                for user in class_users:
                    user_attendance[user['id']] = {
                        day: 0 for day in range(1, last_day + 1)}
                    user_attendance[user['id']]['user_info'] = user

                for user_log_data in log_data:
                    user_id = user_log_data.get('user_id')
                    if user_id not in user_attendance:
                        continue
                    for _, log in user_log_data['logs'].iterrows():
                        event_time = pd.to_datetime(log['timestamp'])
                        if event_time.year == year and event_time.month == month and log['event'] == 'Check-in':
                            user_attendance[user_id][event_time.day] += 1

                exported_data = []
                for user_id, attendance_data in user_attendance.items():
                    user_info_dict = attendance_data['user_info']
                    daily_counts = [attendance_data[day]
                                    for day in range(1, last_day + 1)]
                    days_attended_by_user = sum(
                        1 for count in daily_counts if count > 0)
                    total_present_person_days += days_attended_by_user
                    if user_info_dict.get('gender', '').lower() in ['female', 'ស្រី']:
                        female_present_person_days += days_attended_by_user

                    user_info_export = {
                        'ID': user_id, 'Name': f"{user_info_dict.get('first_name', '')} {user_info_dict.get('last_name', '')}".strip(),
                        'Sex': str(user_info_dict.get('gender', 'N/A')), 'Class': str(user_info_dict.get('class', 'N/A')),
                        'Study Year': str(user_info_dict.get('study_year', DEFAULT_STUDY_YEAR)), 'Total Attendance': sum(daily_counts)
                    }
                    for day in range(1, last_day + 1):
                        count = attendance_data[day]
                        user_info_export[f'{day}'] = count if count > 0 else ''
                    exported_data.append(user_info_export)

                wb = Workbook()
                ws = wb.active
                ws.title = "Monthly Attendance Report"
                ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
                ws.page_setup.paperSize = ws.PAPERSIZE_A4
                ws.page_setup.fitToPage = True
                ws.page_setup.fitToWidth, ws.page_setup.fitToHeight = 1, 0

                thin_border = Border(left=Side(style='thin'), right=Side(
                    style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                columns = ['ល.រ', 'អត្ថលេខ', 'ឈ្មោះ', 'ភេទ', 'ថ្នាក់', 'ឆ្នាំសិក្សា'] + \
                    [f'{day}' for day in range(1, last_day + 1)] + ['សរុប']
                end_col_letter = get_column_letter(len(columns))

                khmer_months = ["មករា", "កុម្ភៈ", "មីនា", "មេសា", "ឧសភា",
                                "មិថុនា", "កក្កដា", "សីហា", "កញ្ញា", "តុលា", "វិច្ឆិកា", "ធ្នូ"]

                # Sheet Header
                ws.merge_cells(f'A1:{end_col_letter}1')
                ws['A1'].value = "ព្រះរាជាណាចក្រកម្ពុជា"
                ws['A1'].font = Font(name='Khmer OS Muol', size=12, bold=True)
                ws['A1'].alignment = Alignment(horizontal='center')
                ws.merge_cells(f'A2:{end_col_letter}2')
                ws['A2'].value = "ជាតិ សាសនា ព្រះមហាក្សត្រ"
                ws['A2'].font = Font(name='Khmer OS Muol', size=11, bold=True)
                ws['A2'].alignment = Alignment(horizontal='center')
                ws.merge_cells('A4:F4')
                ws['A4'].value = "វិទ្យាល័យសម្តេចឪ សម្តេចម៉ែ"
                ws['A4'].font = Font(name='Khmer OS Muol', size=10, bold=True)
                ws.merge_cells(f'A5:{end_col_letter}5')
                ws['A5'].value = f"របាយការណ៍វត្តមានប្រចាំខែ {khmer_months[month - 1]}"
                ws['A5'].font = Font(name='Khmer OS Muol', size=11, bold=True)
                ws['A5'].alignment = Alignment(horizontal='center')
                ws.merge_cells(f'A6:{end_col_letter}6')
                ws['A6'].value = f"ថ្នាក់ទី {selected_class}"
                ws['A6'].font = Font(name='Khmer OS Muol', size=11, bold=True)
                ws['A6'].alignment = Alignment(horizontal='center')

                header_row = 8
                for col_num, header_title in enumerate(columns, 1):
                    cell = ws.cell(
                        row=header_row, column=col_num, value=header_title)
                    cell.font = Font(name='Khmer OS Siemreap', bold=True)
                    cell.alignment = Alignment(
                        horizontal='center', vertical='center')
                    cell.border = thin_border

                for row_idx, user_info in enumerate(sorted(exported_data, key=lambda x: x['ID']), 1):
                    row_data = [row_idx, user_info['ID'], user_info['Name'], user_info['Sex'], user_info['Class'], user_info['Study Year']] + \
                               [user_info.get(f'{day}', '') for day in range(
                                   1, last_day + 1)] + [user_info['Total Attendance']]
                    for j, value in enumerate(row_data):
                        cell = ws.cell(row=header_row + row_idx,
                                       column=j + 1, value=value)
                        cell.font = Font(name='Khmer OS Siemreap')
                        cell.alignment = Alignment(horizontal='center')
                        cell.border = thin_border

                summary_row = header_row + len(exported_data) + 2
                ws[f'A{summary_row}'].value = f"សរុបវត្តមានសិស្ស (គិតជាថ្ងៃ): {total_present_person_days}"
                ws[f'A{summary_row}'].font = Font(
                    name='Khmer OS Siemreap', bold=True)
                ws[f'A{summary_row + 1}'].value = f"សរុបវត្តមានសិស្សស្រី (គិតជាថ្ងៃ): {female_present_person_days}"
                ws[f'A{summary_row + 1}'].font = Font(
                    name='Khmer OS Siemreap', bold=True)

                footer_start_row = summary_row + 2
                lunar_start_col_index = self.signature_col_spinbox.value()
                lunar_end_col_index = lunar_start_col_index + 15
                today = date.today()

                ws.merge_cells(start_row=footer_start_row, start_column=lunar_start_col_index,
                               end_row=footer_start_row, end_column=lunar_end_col_index)
                ws.cell(row=footer_start_row, column=lunar_start_col_index).value = "ធ្វើនៅ ថ្ងៃ.......... ............. ខែ.......... ឆ្នាំ.......... .....ស័ក ព.ស.........."
                ws.cell(row=footer_start_row, column=lunar_start_col_index).font = Font(
                    name='Khmer OS Siemreap')
                ws.cell(row=footer_start_row, column=lunar_start_col_index).alignment = Alignment(
                    horizontal='center')
                ws.merge_cells(start_row=footer_start_row + 1, start_column=lunar_start_col_index,
                               end_row=footer_start_row + 1, end_column=lunar_end_col_index)
                ws.cell(row=footer_start_row + 1,
                        column=lunar_start_col_index).value = f"ត្រូវនឹងថ្ងៃទី {today.day} ខែ {khmer_months[today.month - 1]} ឆ្នាំ {today.year}"
                ws.cell(row=footer_start_row + 1,
                        column=lunar_start_col_index).font = Font(name='Khmer OS Siemreap')
                ws.cell(row=footer_start_row + 1,
                        column=lunar_start_col_index).alignment = Alignment(horizontal='center')
                ws.merge_cells(start_row=footer_start_row + 3, start_column=lunar_start_col_index,
                               end_row=footer_start_row + 3, end_column=lunar_end_col_index)
                ws.cell(row=footer_start_row + 3,
                        column=lunar_start_col_index).value = "ហត្ថលេខាបណ្ណារ័ក្ស"
                ws.cell(row=footer_start_row + 3, column=lunar_start_col_index).font = Font(
                    name='Khmer OS Siemreap', size=10, bold=False)
                ws.cell(row=footer_start_row + 3,
                        column=lunar_start_col_index).alignment = Alignment(horizontal='center')

                # Column Widths
                for col_idx, column_cells in enumerate(ws.columns, 1):
                    col_letter = get_column_letter(col_idx)
                    if col_idx == 1:
                        ws.column_dimensions[col_letter].width = 5
                        continue
                    if col_idx - 1 < len(columns) and columns[col_idx - 1].isdigit():
                        ws.column_dimensions[col_letter].width = 4
                        continue
                    max_length = max(
                        (len(str(cell.value)) for cell in column_cells if cell.value is not None), default=0)
                    ws.column_dimensions[col_letter].width = max_length + 2

                filename = f"monthly_class_{selected_class}_attendance_{year}{month:02d}.xlsx"
                filepath = os.path.join(output_folder, filename)
                wb.save(filepath)
                success_count += 1

            except PermissionError:
                logging.error(
                    f"Permission denied. Could not save file for {month_name} to {filepath}. Please ensure the file is not open elsewhere.")
                failed_months.append(month_name)
            except Exception as e:
                logging.error(
                    f"Failed to export {month_name} for class {selected_class}: {e}")
                failed_months.append(month_name)

        progress.setValue(len(months))
        if success_count > 0:
            message = f"Successfully exported {success_count} monthly reports to:\n{output_folder}"
            if failed_months:
                message += f"\nFailed to export: {', '.join(failed_months)}"
            QMessageBox.information(self, "Export Complete", message)
        else:
            QMessageBox.warning(
                self, "Export Failed", "Could not export any of the selected monthly reports.")


class SkippedUsersDialog(QDialog):
    def __init__(self, skipped_users, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Import Warnings")
        self.setMinimumSize(400, 300)
        layout = QVBoxLayout(self)

        header = QLabel(
            f"<b>{len(skipped_users)}</b> user(s) were skipped because their ID already exists:")
        header.setWordWrap(True)
        layout.addWidget(header)

        list_widget = QListWidget()
        list_widget.addItems(skipped_users)
        layout.addWidget(list_widget)

        ok_button = QPushButton("OK")
        ok_button.clicked.connect(self.accept)
        layout.addWidget(ok_button, alignment=Qt.AlignCenter)


# Main Entry Point
if __name__ == "__main__":
    try:
        app = QApplication(sys.argv)
        user_manager = UserManager(USER_DATA_FILE)
        log_manager = LogManager(LOG_DATABASE_FILE)
        window = QRCodeApp(user_manager, log_manager)
        window.show()
        sys.exit(app.exec_())
    except Exception as e:
        logging.critical(f"Application crashed: {e}\n{traceback.format_exc()}")
        traceback.print_exc()
